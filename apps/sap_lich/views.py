"""
Views for Sap Lich (Scheduling) app
Provides admin interface for LLM and algorithm-based scheduling
"""

import json
import logging
import random
import time
from datetime import datetime, timedelta
from functools import wraps
from django.contrib import admin
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponseForbidden
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q, Prefetch
from apps.scheduling.models import (
    DotXep, ThoiKhoaBieu, GiangVien, PhongHoc, 
    TimeSlot, KhungTG, PhanCong, LopMonHoc, MonHoc,
    NguyenVong, GVDayMon, TKBLog
)

logger = logging.getLogger(__name__)

def get_user_role_info(user):
    """
    Xác định role và thông tin liên quan của user
    Returns: {
        'role': 'admin' | 'truong_khoa' | 'truong_bo_mon' | 'giang_vien',
        'ma_khoa': str | None,
        'ma_bo_mon': str | None,
        'ma_gv': str | None,
    }
    """
    if user.is_superuser:
        return {'role': 'admin', 'ma_khoa': None, 'ma_bo_mon': None, 'ma_gv': None}
    
    # Lấy groups của user
    groups = user.groups.values_list('name', flat=True)
    
    # Tìm GiangVien theo username (giả định username = ma_gv)
    try:
        giang_vien = GiangVien.objects.select_related('ma_bo_mon', 'ma_bo_mon__ma_khoa').get(ma_gv=user.username)
        ma_gv = giang_vien.ma_gv
        ma_bo_mon = giang_vien.ma_bo_mon.ma_bo_mon if giang_vien.ma_bo_mon else None
        ma_khoa = giang_vien.ma_bo_mon.ma_khoa.ma_khoa if giang_vien.ma_bo_mon and giang_vien.ma_bo_mon.ma_khoa else None
    except GiangVien.DoesNotExist:
        ma_gv = None
        ma_bo_mon = None
        ma_khoa = None
    
    # Xác định role dựa trên group
    if 'Trưởng Khoa' in groups:
        return {'role': 'truong_khoa', 'ma_khoa': ma_khoa, 'ma_bo_mon': None, 'ma_gv': ma_gv}
    elif 'Trưởng Bộ Môn' in groups:
        return {'role': 'truong_bo_mon', 'ma_khoa': ma_khoa, 'ma_bo_mon': ma_bo_mon, 'ma_gv': ma_gv}
    elif 'Giảng Viên' in groups:
        return {'role': 'giang_vien', 'ma_khoa': None, 'ma_bo_mon': None, 'ma_gv': ma_gv}
    else:
        # Mặc định là giáo viên nếu không có group
        return {'role': 'giang_vien', 'ma_khoa': None, 'ma_bo_mon': None, 'ma_gv': ma_gv}


def require_role(*allowed_roles):
    """
    Decorator để kiểm tra role của user
    Usage: @require_role('admin', 'truong_khoa')
    """
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return HttpResponseForbidden("Bạn cần đăng nhập để truy cập trang này")
            
            user_role_info = get_user_role_info(request.user)
            user_role = user_role_info['role']
            
            if user_role not in allowed_roles:
                return HttpResponseForbidden(f"Bạn không có quyền truy cập trang này. Yêu cầu role: {', '.join(allowed_roles)}")
            
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


def landing_page_view(request):
    """
    Landing page - redirect users to appropriate page based on their role
    """
    if not request.user.is_authenticated:
        # Redirect to login page
        from django.contrib.auth.views import redirect_to_login
        return redirect_to_login(request.get_full_path())
    
    role_info = get_user_role_info(request.user)
    user_role = role_info['role']
    ma_gv = role_info['ma_gv'] or request.user.username
    
    # Redirect based on role
    if user_role == 'admin':
        return redirect('/admin/sap_lich/thoikhoabieu/')
    elif user_role == 'truong_khoa':
        return redirect(f'/truong-khoa/{ma_gv}/xem-tkb/')
    elif user_role == 'truong_bo_mon':
        return redirect(f'/truong-bo-mon/{ma_gv}/xem-tkb/')
    else:  # giang_vien or default
        return redirect(f'/giang-vien/{ma_gv}/xem-tkb/')


@csrf_exempt
@require_http_methods(["GET"])
def algo_scheduler_get_stats_api(request):
    """
    API endpoint để lấy thống kê đầu vào của đợt xếp lịch
    
    Query params:
        ma_dot: Mã đợt xếp lịch
    
    Returns:
        {
            "status": "success",
            "stats": {
                "phan_cong": 150,
                "lop_mon_hoc": 120,
                "giang_vien": 45,
                "phong_hoc": 30,
                "mon_hoc": 60,
                "time_slots": 50,
                "tkb_existing": 0
            }
        }
    """
    try:
        ma_dot = request.GET.get('ma_dot')
        
        if not ma_dot:
            return JsonResponse({
                'status': 'error',
                'message': 'Vui lòng cung cấp ma_dot'
            }, status=400)
        
        # Get the scheduling period
        try:
            dot_xep = DotXep.objects.get(ma_dot=ma_dot)
        except DotXep.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': f'Không tìm thấy đợt xếp với mã: {ma_dot}'
            }, status=404)
        
        # Get statistics - Lấy theo đợt xếp cụ thể
        phan_cong_count = PhanCong.objects.filter(ma_dot=dot_xep).count()
        
        # Lớp môn học trong đợt (từ PhanCong)
        lop_mon_hoc_ids = PhanCong.objects.filter(ma_dot=dot_xep).values_list('ma_lop', flat=True).distinct()
        lop_mon_hoc_count = len(lop_mon_hoc_ids)
        
        # Giảng viên tham gia trong đợt (từ PhanCong)
        giang_vien_ids = PhanCong.objects.filter(ma_dot=dot_xep).values_list('ma_gv', flat=True).distinct()
        giang_vien_count = len([gv for gv in giang_vien_ids if gv])
        
        # Môn học trong đợt (từ LopMonHoc của PhanCong)
        from apps.scheduling.models import LopMonHoc
        mon_hoc_ids = LopMonHoc.objects.filter(ma_lop__in=lop_mon_hoc_ids).values_list('ma_mon_hoc', flat=True).distinct()
        mon_hoc_count = len(mon_hoc_ids)
        
        # Phòng học - tổng số có thể dùng (toàn bộ vì phòng không thuộc đợt)
        phong_hoc_count = PhongHoc.objects.count()
        
        # Phòng theo loại (dùng đúng giá trị trong DB: "Lý thuyết", "Thực hành")
        phong_ly_thuyet = PhongHoc.objects.filter(loai_phong='Lý thuyết').count()
        phong_thuc_hanh = PhongHoc.objects.filter(loai_phong='Thực hành').count()
        
        # Time slots
        time_slots_count = TimeSlot.objects.count()
        
        # Nguyện vọng trong đợt
        nguyen_vong_count = NguyenVong.objects.filter(ma_dot=dot_xep).count()
        
        return JsonResponse({
            'status': 'success',
            'ma_dot': ma_dot,
            'ten_dot': dot_xep.ten_dot,
            'stats': {
                'phan_cong': phan_cong_count,
                'lop_mon_hoc': lop_mon_hoc_count,
                'giang_vien': giang_vien_count,
                'phong_hoc': phong_hoc_count,
                'phong_ly_thuyet': phong_ly_thuyet,
                'phong_thuc_hanh': phong_thuc_hanh,
                'mon_hoc': mon_hoc_count,
                'time_slots': time_slots_count,
                'nguyen_vong': nguyen_vong_count
            }
        })
        
    except Exception as e:
        logger.error(f"Error getting stats: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': f'Lỗi: {str(e)}'
        }, status=500)


@require_role('admin')
def llm_scheduler_view(request, ma_gv=None):
    """Admin view for LLM Chatbot Assistant - CHỈ ADMIN"""
    role_info = get_user_role_info(request.user)
    
    if role_info['role'] != 'admin' and not ma_gv:
        # Nếu user truy cập URL cũ (/admin/sap_lich/...), redirect sang URL mới
        ma_gv_current = role_info['ma_gv'] or request.user.username
        if 'truong_khoa' in role_info['role']:
            return redirect(f'/truong-khoa/{ma_gv_current}/xem-tkb/')
        elif 'truong_bo_mon' in role_info['role']:
            return redirect(f'/truong-bo-mon/{ma_gv_current}/xem-tkb/')
        else:
            return redirect(f'/giang-vien/{ma_gv_current}/xem-tkb/')
    
    if role_info['role'] != 'admin':
        return HttpResponseForbidden("Chỉ Admin mới có quyền truy cập Chat bot hỗ trợ")
    
    try:
        periods = list(DotXep.objects.all().values('ma_dot', 'ten_dot', 'trang_thai'))
    except Exception:
        periods = []
    
    # Get admin site context with proper breadcrumb info
    context = {
        **admin.site.each_context(request),
        'periods': periods,
        'title': 'Trợ lý AI - Hỏi đáp Lịch học',
        'site_title': admin.site.site_title,
        'site_header': admin.site.site_header,
        'has_permission': True,
        'is_nav_sidebar_enabled': True,
        'app_label': 'sap_lich',
        'opts': {
            'app_label': 'sap_lich',
            'model_name': 'saplich',
            'verbose_name_plural': 'Sắp lịch',
        },
        'current_time': datetime.now().strftime('%H:%M'),
        'user_role': role_info['role'],
        'segment': ['sap-lich-llm'],
    }
    return render(request, 'admin/llm_scheduler.html', context)


@require_role('admin')
def algo_scheduler_view(request, ma_gv=None):
    """Admin view for algorithm-based scheduler - CHỈ ADMIN"""
    role_info = get_user_role_info(request.user)
    
    if role_info['role'] != 'admin' and not ma_gv:
        # Nếu user truy cập URL cũ (/admin/sap_lich/...), redirect sang URL mới
        ma_gv_current = role_info['ma_gv'] or request.user.username
        if 'truong_khoa' in role_info['role']:
            return redirect(f'/truong-khoa/{ma_gv_current}/xem-tkb/')
        elif 'truong_bo_mon' in role_info['role']:
            return redirect(f'/truong-bo-mon/{ma_gv_current}/xem-tkb/')
        else:
            return redirect(f'/giang-vien/{ma_gv_current}/xem-tkb/')
    
    if role_info['role'] != 'admin':
        return HttpResponseForbidden("Chỉ Admin mới có quyền sắp lịch bằng thuật toán")
    
    try:
        periods = list(DotXep.objects.all().values('ma_dot', 'ten_dot', 'trang_thai'))
    except Exception:
        periods = []
    
    # Get admin site context with proper breadcrumb info
    context = {
        **admin.site.each_context(request),
        'periods': periods,
        'title': 'Sắp lịch bằng thuật toán',
        'site_title': admin.site.site_title,
        'site_header': admin.site.site_header,
        'has_permission': True,
        'is_nav_sidebar_enabled': True,
        'app_label': 'sap_lich',
        'opts': {
            'app_label': 'sap_lich',
            'model_name': 'saplich',
            'verbose_name_plural': 'Sắp lịch',
        },
        'user_role': role_info['role'],
        'segment': ['sap-lich-algo'],
    }
    return render(request, 'admin/algo_scheduler.html', context)


@require_role('admin')
@csrf_exempt
@require_http_methods(["POST"])
def algo_scheduler_run_api(request):
    """
    API endpoint để chạy thuật toán xếp lịch - CHỈ ADMIN
    
    Expected POST data:
    {
        "ma_dot": "2025-2026_HK1",
        "strategy": "TS",  // "TS" (Tabu Search) hoặc "SA" (Simulated Annealing)
        "init_method": "greedy-cprop",  // "greedy-cprop" hoặc "random-repair"
        "time_limit": 180,  // seconds (default 180s = 3 phút)
        "seed": 42,  // optional, random seed
        "save_to_db": true  // optional, lưu vào ThoiKhoaBieu hay không
    }
    
    Returns:
    {
        "status": "success",
        "ma_dot": "2025-2026_HK1",
        "initial_cost": 145,
        "final_cost": 89,
        "improvement": 56,
        "improvement_percent": 38.6,
        "time_elapsed": 180.5,
        "breakdown": {
            "room_capacity": 0,
            "min_working_days": 0,
            "curriculum_compactness": 45,
            "lecture_consecutiveness": 0,
            "room_stability": 0,
            "teacher_preferences": 44
        },
        "sol_file": "/path/to/solution.sol",
        "saved_to_db": true,
        "message": "Xếp lịch thành công!"
    }
    """
    try:
        from apps.scheduling.algorithms.algorithms_runner import AlgorithmRunner
        
        data = json.loads(request.body)
        ma_dot = data.get('ma_dot')
        strategy = data.get('strategy', 'TS').upper()
        init_method = data.get('init_method', 'greedy-cprop')
        time_limit = float(data.get('time_limit', 180))
        seed = data.get('seed', 42)
        save_to_db = data.get('save_to_db', True)

        # Validation
        if not ma_dot:
            return JsonResponse({
                'status': 'error',
                'message': 'Vui lòng cung cấp ma_dot'
            }, status=400)

        if strategy not in ['TS', 'SA']:
            return JsonResponse({
                'status': 'error',
                'message': 'Strategy không hợp lệ. Phải là "TS" hoặc "SA"'
            }, status=400)

        if init_method not in ['greedy-cprop', 'random-repair']:
            return JsonResponse({
                'status': 'error',
                'message': 'Init method không hợp lệ. Phải là "greedy-cprop" hoặc "random-repair"'
            }, status=400)

        logger.info(f"🚀 Bắt đầu xếp lịch cho {ma_dot}")
        logger.info(f"   Strategy: {strategy}, Init: {init_method}, Time: {time_limit}s, Seed: {seed}")

        # Step 1: Initialize runner
        runner = AlgorithmRunner(ma_dot=ma_dot, seed=seed)

        # Step 2: Prepare data (export DB to CTT)
        logger.info("📊 Step 1: Chuẩn bị dữ liệu (export DB sang CTT)")
        if not runner.prepare_data():
            return JsonResponse({
                'status': 'error',
                'message': 'Không thể chuẩn bị dữ liệu. Kiểm tra xem DotXep có tồn tại và có dữ liệu hợp lệ không.'
            }, status=400)

        # Step 3: Run optimization
        logger.info("🔧 Step 2: Chạy thuật toán optimization")
        result = runner.run_optimization(
            strategy=strategy,
            init_method=init_method,
            time_limit=time_limit
        )

        if not result or not result.get('success'):
            error_msg = result.get('error', 'Thuật toán thất bại') if result else 'Lỗi không xác định'
            logger.error(f"❌ Optimization failed: {error_msg}")
            return JsonResponse({
                'status': 'error',
                'message': error_msg
            }, status=500)

        # Step 4: Save to database (nếu requested)
        if save_to_db:
            logger.info("💾 Step 3: Lưu kết quả vào database")
            
            # Reconstruct assignments from formatted result
            assignments = {}
            for lecture_id_str, assignment_data in result.get('assignments', {}).items():
                lecture_id = int(lecture_id_str)
                period = assignment_data['period_absolute']
                
                # Find room_idx from room_id
                room_id = assignment_data['room_id']
                room_idx = None
                for idx, room in enumerate(runner.instance.rooms):
                    if room.id == room_id:
                        room_idx = idx
                        break
                
                if room_idx is not None:
                    assignments[lecture_id] = (period, room_idx)
            
            saved = runner.save_to_database(assignments)
            result['saved_to_db'] = saved
            
            if not saved:
                logger.warning("⚠️  Lưu vào database thất bại, nhưng optimization thành công")
                result['warning'] = 'Lưu vào database thất bại'
        else:
            result['saved_to_db'] = False

        # Format response
        logger.info(f"✅ Xếp lịch hoàn tất!")
        logger.info(f"   Initial cost: {result['initial_cost']}")
        logger.info(f"   Final cost: {result['final_cost']}")
        logger.info(f"   Improvement: {result['improvement']} ({result['improvement_percent']:.1f}%)")
        logger.info(f"   Teacher preferences: {result['breakdown']['teacher_preferences']} violations")

        # Convert to JsonResponse format
        response = {
            'status': 'success',
            'ma_dot': result['ma_dot'],
            'initial_cost': result['initial_cost'],
            'final_cost': result['final_cost'],
            'improvement': result['improvement'],
            'improvement_percent': round(result['improvement_percent'], 2),
            'time_elapsed': round(result['time_elapsed'], 2),
            'breakdown': result['breakdown'],
            'sol_file': result['sol_file'],
            'saved_to_db': result['saved_to_db'],
            'message': f'Xếp lịch thành công! Cost giảm từ {result["initial_cost"]} xuống {result["final_cost"]} ({result["improvement_percent"]:.1f}%)',
            'details': {
                'strategy': strategy,
                'init_method': init_method,
                'seed': seed,
                'lectures_scheduled': len(result.get('assignments', {}))
            }
        }

        if 'warning' in result:
            response['warning'] = result['warning']

        return JsonResponse(response)

    except json.JSONDecodeError:
        logger.error("JSON không hợp lệ")
        return JsonResponse({
            'status': 'error',
            'message': 'JSON không hợp lệ'
        }, status=400)
    except Exception as e:
        logger.exception(f"Lỗi API: {e}")
        return JsonResponse({
            'status': 'error',
            'message': f'Lỗi: {str(e)}'
        }, status=500)


@require_http_methods(["GET"])
def algo_scheduler_view_result_api(request):
    """
    API endpoint để xem kết quả thời khóa biểu đã được lưu vào database
    
    Expected GET parameters:
    - ma_dot: Mã đợt xếp lịch
    
    Returns:
    {
        "status": "success",
        "ma_dot": "2025-2026_HK1",
        "ten_dot": "Học kỳ 1 năm 2025-2026",
        "total_schedules": 150,
        "schedules": [
            {
                "ma_lop": "CTTT01",
                "ten_lop": "Cấu trúc dữ liệu",
                "ma_gv": "GV001",
                "ten_gv": "Nguyễn Văn A",
                "ma_phong": "A101",
                "thu": 2,
                "ca": 1,
                "tuan_hoc": 1
            },
            ...
        ]
    }
    """
    try:
        ma_dot = request.GET.get('ma_dot')
        
        if not ma_dot:
            return JsonResponse({
                'status': 'error',
                'message': 'Vui lòng cung cấp ma_dot'
            }, status=400)
        
        # Kiểm tra đợt xếp có tồn tại không
        try:
            dot_xep = DotXep.objects.get(ma_dot=ma_dot)
        except DotXep.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': f'Không tìm thấy đợt xếp {ma_dot}'
            }, status=404)
        
        # Lấy tất cả thời khóa biểu của đợt
        from apps.scheduling.models import ThoiKhoaBieu, PhanCong
        
        tkb_list = ThoiKhoaBieu.objects.filter(
            ma_dot=dot_xep,
            is_deleted=False
        ).select_related(
            'ma_lop',
            'ma_lop__ma_mon_hoc',
            'ma_phong',
            'time_slot_id__ca'
        ).order_by('time_slot_id__thu', 'time_slot_id__ca__ma_khung_gio')
        
        # Lấy mapping từ ma_lop sang giảng viên qua PhanCong
        lop_to_gv = {}
        for pc in PhanCong.objects.filter(ma_dot=dot_xep).select_related('ma_lop', 'ma_gv'):
            lop_to_gv[pc.ma_lop.ma_lop] = pc.ma_gv
        
        # Format kết quả
        schedules = []
        for tkb in tkb_list:
            # Lấy thông tin giảng viên từ mapping
            gv = lop_to_gv.get(tkb.ma_lop.ma_lop)
            
            schedules.append({
                'id': tkb.ma_tkb,
                'ma_lop': tkb.ma_lop.ma_lop,
                'ten_lop': f"{tkb.ma_lop.ma_mon_hoc.ten_mon_hoc} (Nhóm {tkb.ma_lop.nhom_mh})",
                'ma_mon': tkb.ma_lop.ma_mon_hoc.ma_mon_hoc,
                'ten_mon': tkb.ma_lop.ma_mon_hoc.ten_mon_hoc,
                'ma_gv': gv.ma_gv if gv else 'N/A',
                'ten_gv': gv.ten_gv if gv else 'Chưa phân công',
                'ma_phong': tkb.ma_phong.ma_phong if tkb.ma_phong else 'N/A',
                'suc_chua': tkb.ma_phong.suc_chua if tkb.ma_phong and tkb.ma_phong.suc_chua else 0,
                'loai_phong': tkb.ma_phong.loai_phong if tkb.ma_phong else 'N/A',
                'thu': tkb.time_slot_id.thu,
                'ca': tkb.time_slot_id.ca.ma_khung_gio,
                'gio_bat_dau': str(tkb.time_slot_id.ca.gio_bat_dau),
                'gio_ket_thuc': str(tkb.time_slot_id.ca.gio_ket_thuc),
                'tuan_hoc': tkb.tuan_hoc if tkb.tuan_hoc else '1',
            })
        
        logger.info(f"Retrieved {len(schedules)} schedules for {ma_dot}")
        
        # Chạy validator để lấy breakdown costs từ file .sol đã lưu
        breakdown = None
        initial_cost = None
        final_cost = None
        
        try:
            from pathlib import Path
            import subprocess
            import re
            from django.conf import settings
            
            # Path to .ctt and .sol files
            ctt_file = Path(settings.BASE_DIR) / 'output' / 'test_web_algo' / 'ctt_files' / f'{ma_dot}.ctt'
            sol_file = Path(settings.BASE_DIR) / 'output' / 'test_web_algo' / f'solution_{ma_dot}.sol'
            
            if ctt_file.exists() and sol_file.exists():
                # Run validator
                result = subprocess.run(
                    ['python', 'apps/scheduling/utils/validator.py', str(ctt_file), str(sol_file)],
                    capture_output=True,
                    text=True,
                    cwd=settings.BASE_DIR
                )
                
                if result.returncode == 0:
                    output = result.stdout
                    
                    # Parse costs from validator output
                    breakdown = {}
                    cost_patterns = {
                        'min_working_days': r'Cost of MinWorkingDays \(soft\)\s*:\s*(\d+)',
                        'curriculum_compactness': r'Cost of CurriculumCompactness \(soft\)\s*:\s*(\d+)',
                        'lecture_consecutiveness': r'Cost of LectureConsecutiveness \(soft\)\s*:\s*(\d+)',
                        'room_stability': r'Cost of RoomStability \(soft\)\s*:\s*(\d+)',
                        'teacher_lecture_consolidation': r'Cost of TeacherLectureConsolidation \(soft - extended\)\s*:\s*(\d+)',
                        'teacher_working_days': r'Cost of TeacherWorkingDays \(soft - extended\)\s*:\s*(\d+)',
                        'teacher_preferences': r'Cost of TeacherPreferences \(soft - extended\)\s*:\s*(\d+)',
                        'room_capacity': r'Cost of RoomCapacity \(soft\)\s*:\s*(\d+)',
                    }
                    
                    for key, pattern in cost_patterns.items():
                        match = re.search(pattern, output)
                        if match:
                            breakdown[key] = int(match.group(1))
                    
                    # Parse total cost
                    total_match = re.search(r'Total Cost = (\d+)', output)
                    if total_match:
                        final_cost = int(total_match.group(1))
                        initial_cost = final_cost  # Không có initial cost khi load từ DB
                    
                    logger.info(f"Validator breakdown: {breakdown}")
                else:
                    logger.warning(f"Validator failed with code {result.returncode}: {result.stderr}")
        except Exception as e:
            logger.warning(f"Could not run validator: {e}")
        
        response = {
            'status': 'success',
            'ma_dot': ma_dot,
            'ten_dot': dot_xep.ten_dot,
            'total_schedules': len(schedules),
            'schedules': schedules
        }
        
        # Add breakdown if available
        if breakdown:
            response['breakdown'] = breakdown
            response['final_cost'] = final_cost
            response['initial_cost'] = initial_cost
        
        return JsonResponse(response)
    
    except Exception as e:
        logger.exception(f"Lỗi khi xem kết quả: {e}")
        return JsonResponse({
            'status': 'error',
            'message': f'Lỗi: {str(e)}'
        }, status=500)


@require_http_methods(["GET"])
def algo_scheduler_get_weights_api(request):
    """
    API endpoint để lấy trọng số của các ràng buộc cho đợt xếp lịch
    
    Expected GET parameters:
    - ma_dot: Mã đợt xếp lịch (optional)
    
    Returns:
    {
        "status": "success",
        "weights": {
            "RBM-001": {"name": "...", "weight": 2.5, "source": "dot"},
            "RBM-002": {"name": "...", "weight": 1.0, "source": "global"},
            ...
        }
    }
    """
    try:
        from apps.scheduling.algorithms.weight_loader import WeightLoader
        from apps.scheduling.models import RangBuocMem, RangBuocTrongDot
        
        ma_dot = request.GET.get('ma_dot')
        
        # Load weights using WeightLoader (3-tier priority)
        weights = WeightLoader.load_weights(ma_dot)
        
        # Get mapping from RBM codes to friendly names
        rang_buoc_map = {}
        for rb in RangBuocMem.objects.all():
            rang_buoc_map[rb.ma_rang_buoc] = rb.ten_rang_buoc
        
        # Get dot-specific overrides if ma_dot provided
        dot_overrides = set()
        if ma_dot:
            dot_overrides = set(
                RangBuocTrongDot.objects.filter(ma_dot=ma_dot)
                .values_list('ma_rang_buoc__ma_rang_buoc', flat=True)
            )
        
        # Map internal keys back to RBM codes with weight values
        # Reverse lookup from CONSTRAINT_MAPPING
        from apps.scheduling.algorithms.weight_loader import CONSTRAINT_MAPPING
        rbm_weights = {}
        
        for rbm_code, internal_key in CONSTRAINT_MAPPING.items():
            if internal_key in weights:
                source = 'dot' if rbm_code in dot_overrides else 'global'
                rbm_weights[rbm_code] = {
                    'name': rang_buoc_map.get(rbm_code, internal_key),
                    'weight': weights[internal_key],
                    'source': source
                }
        
        return JsonResponse({
            'status': 'success',
            'weights': rbm_weights
        })
    
    except Exception as e:
        logger.exception(f"Lỗi khi lấy weights: {e}")
        return JsonResponse({
            'status': 'error',
            'message': f'Lỗi: {str(e)}'
        }, status=500)


@require_http_methods(["GET"])
def algo_scheduler_export_excel_api(request):
    """
    API endpoint để xuất thời khóa biểu ra file Excel
    
    Expected GET parameters:
    - ma_dot: Mã đợt xếp lịch
    
    Returns:
        Excel file download
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    try:
        ma_dot = request.GET.get('ma_dot')
        
        if not ma_dot:
            return JsonResponse({
                'status': 'error',
                'message': 'Vui lòng cung cấp ma_dot'
            }, status=400)
        
        # Kiểm tra đợt xếp có tồn tại không
        try:
            dot_xep = DotXep.objects.get(ma_dot=ma_dot)
        except DotXep.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': f'Không tìm thấy đợt xếp {ma_dot}'
            }, status=404)
        
        # Lấy tất cả thời khóa biểu của đợt
        tkb_list = ThoiKhoaBieu.objects.filter(
            ma_dot=dot_xep
        ).select_related(
            'ma_lop',
            'ma_lop__ma_mon_hoc',
            'ma_phong',
            'time_slot_id__ca'
        ).order_by('time_slot_id__thu', 'time_slot_id__ca__ma_khung_gio')
        
        # Lấy mapping từ ma_lop sang giảng viên qua PhanCong
        lop_to_gv = {}
        for pc in PhanCong.objects.filter(ma_dot=dot_xep).select_related('ma_lop', 'ma_gv'):
            lop_to_gv[pc.ma_lop.ma_lop] = pc.ma_gv
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"TKB_{ma_dot}"
        
        # Styling
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['STT', 'Mã Lớp', 'Tên Môn Học', 'Nhóm', 'Mã GV', 'Tên GV', 
                   'Mã Phòng', 'Loại Phòng', 'Sức Chứa', 'Thứ', 'Ca', 'Giờ BĐ', 'Giờ KT', 'Tuần Học']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Mapping thứ
        day_map = {
            2: 'Thứ 2', 3: 'Thứ 3', 4: 'Thứ 4', 5: 'Thứ 5',
            6: 'Thứ 6', 7: 'Thứ 7', 8: 'Chủ Nhật'
        }
        
        # Write data
        for row_num, tkb in enumerate(tkb_list, 2):
            gv = lop_to_gv.get(tkb.ma_lop.ma_lop)
            
            row_data = [
                row_num - 1,  # STT
                tkb.ma_lop.ma_lop,
                tkb.ma_lop.ma_mon_hoc.ten_mon_hoc,
                tkb.ma_lop.nhom_mh,
                gv.ma_gv if gv else 'N/A',
                gv.ten_gv if gv else 'Chưa phân công',
                tkb.ma_phong.ma_phong if tkb.ma_phong else 'N/A',
                tkb.ma_phong.loai_phong if tkb.ma_phong else 'N/A',
                tkb.ma_phong.suc_chua if tkb.ma_phong else 0,
                day_map.get(tkb.time_slot_id.thu, tkb.time_slot_id.thu),
                f"Ca {tkb.time_slot_id.ca.ma_khung_gio}",
                str(tkb.time_slot_id.ca.gio_bat_dau),
                str(tkb.time_slot_id.ca.gio_ket_thuc),
                tkb.tuan_hoc if tkb.tuan_hoc else '1-15'
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = str(value) if value is not None else ''
                cell.border = border
                cell.alignment = Alignment(vertical='center')
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = len(headers[col_num - 1])
            
            for row_num in range(2, min(102, ws.max_row + 1)):
                cell_value = ws[f'{column_letter}{row_num}'].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Create response
        from django.http import HttpResponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'TKB_{ma_dot}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        logger.info(f"Exported {tkb_list.count()} schedules for {ma_dot} to Excel")
        return response
        
    except Exception as e:
        logger.exception(f"Lỗi khi xuất Excel: {e}")
        return JsonResponse({
            'status': 'error',
            'message': f'Lỗi: {str(e)}'
        }, status=500)


def thoikhoabieu_view(request, ma_gv=None):
    """
    View hiển thị thời khóa biểu với nhiều góc nhìn và dạng hiển thị
    - Góc nhìn: theo giáo viên, theo phòng
    - Dạng hiển thị: tổng quát (tất cả tuần), chi tiết (theo tuần)
    
    URL Parameters:
        ma_gv: Mã giảng viên từ URL (dùng cho các role không phải admin)
    
    Phân quyền:
        - Admin: Xem toàn bộ TKB
        - Trưởng Khoa: Xem TKB khoa mình (filter theo ma_khoa)
        - Trưởng Bộ Môn: Xem TKB bộ môn mình (filter theo ma_bo_mon)
        - Giảng Viên: Xem TKB cá nhân (filter theo ma_gv)
    """
    # Kiểm tra authentication
    if not request.user.is_authenticated:
        return HttpResponseForbidden("Bạn cần đăng nhập để xem thời khóa biểu")
    
    # Lấy thông tin phân quyền
    role_info = get_user_role_info(request.user)
    user_role = role_info['role']
    
    # Redirect non-admin users từ URL cũ sang URL mới
    if user_role != 'admin' and not ma_gv:
        # Nếu user truy cập URL cũ (/admin/sap_lich/thoikhoabieu/), redirect sang URL mới
        ma_gv_current = role_info['ma_gv'] or request.user.username
        if user_role == 'truong_khoa':
            return redirect(f'/truong-khoa/{ma_gv_current}/xem-tkb/')
        elif user_role == 'truong_bo_mon':
            return redirect(f'/truong-bo-mon/{ma_gv_current}/xem-tkb/')
        else:
            return redirect(f'/giang-vien/{ma_gv_current}/xem-tkb/')
    
    # Validate ma_gv trong URL với user hiện tại
    if ma_gv:
        # Nếu không phải admin, phải check ma_gv khớp với user
        if user_role != 'admin':
            # Trưởng Khoa, Trưởng Bộ Môn, Giảng Viên phải match với ma_gv của họ
            if ma_gv != role_info['ma_gv']:
                return HttpResponseForbidden("Bạn không có quyền xem thời khóa biểu của người khác")
    
    # Lấy các tham số từ request
    view_type = request.GET.get('view', 'teacher')  # 'teacher' hoặc 'room'
    display_mode = request.GET.get('mode', 'general')  # 'general' hoặc 'weekly'
    week_number = int(request.GET.get('week', 1))  # Tuần hiện tại (1-15)
    ma_dot = request.GET.get('ma_dot', '')  # Đợt xếp lịch
    ma_khoa = request.GET.get('ma_khoa', '')  # Khoa
    ma_bo_mon = request.GET.get('ma_bo_mon', '')  # Bộ môn
    selected_id = request.GET.get('id', '')  # Mã GV hoặc mã phòng
    
    # Áp dụng ràng buộc theo phân quyền
    if user_role == 'truong_khoa':
        # Trưởng khoa chỉ thấy khoa mình
        ma_khoa = role_info['ma_khoa']
    elif user_role == 'truong_bo_mon':
        # Trưởng bộ môn chỉ thấy bộ môn mình
        ma_khoa = role_info['ma_khoa']
        ma_bo_mon = role_info['ma_bo_mon']
    elif user_role == 'giang_vien':
        # Giáo viên chỉ thấy TKB của mình
        view_type = 'teacher'
        selected_id = role_info['ma_gv']
        if role_info['ma_khoa']:
            ma_khoa = role_info['ma_khoa']
        if role_info['ma_bo_mon']:
            ma_bo_mon = role_info['ma_bo_mon']
    
    # Lấy danh sách đợt xếp lịch
    dot_list = DotXep.objects.all().order_by('-ma_dot')
    
    # Lấy danh sách khoa (filter theo quyền)
    from apps.scheduling.models import Khoa, BoMon
    if user_role == 'truong_khoa':
        khoa_list = Khoa.objects.filter(ma_khoa=role_info['ma_khoa']).order_by('ma_khoa')
    elif user_role in ['truong_bo_mon', 'giang_vien'] and role_info['ma_khoa']:
        khoa_list = Khoa.objects.filter(ma_khoa=role_info['ma_khoa']).order_by('ma_khoa')
    else:
        khoa_list = Khoa.objects.all().order_by('ma_khoa')
    
    # Lấy danh sách bộ môn theo khoa đã chọn (filter theo quyền)
    bo_mon_list = []
    if ma_khoa:
        if user_role == 'truong_bo_mon':
            bo_mon_list = BoMon.objects.filter(
                ma_khoa__ma_khoa=ma_khoa,
                ma_bo_mon=role_info['ma_bo_mon']
            ).order_by('ma_bo_mon')
        elif user_role == 'giang_vien' and role_info['ma_bo_mon']:
            bo_mon_list = BoMon.objects.filter(
                ma_khoa__ma_khoa=ma_khoa,
                ma_bo_mon=role_info['ma_bo_mon']
            ).order_by('ma_bo_mon')
        else:
            bo_mon_list = BoMon.objects.filter(ma_khoa__ma_khoa=ma_khoa).order_by('ma_bo_mon')
    
    # Nếu không có ma_dot, lấy đợt mới nhất
    if not ma_dot and dot_list.exists():
        ma_dot = dot_list.first().ma_dot
    
    # Khởi tạo context
    context = {
        **admin.site.each_context(request),
        'title': 'Thời Khóa Biểu',
        'view_type': view_type,
        'display_mode': display_mode,
        'week_number': week_number,
        'ma_dot': ma_dot,
        'ma_khoa': ma_khoa,
        'ma_bo_mon': ma_bo_mon,
        'selected_id': selected_id,
        'dot_list': dot_list,
        'khoa_list': khoa_list,
        'bo_mon_list': bo_mon_list,
        'weeks': range(1, 16),  # 15 tuần
        'user_role': user_role,  # Thêm role vào context
        'role_info': role_info,  # Thêm role_info vào context
        'app_label': 'sap_lich',
        'opts': {
            'app_label': 'sap_lich',
            'model_name': 'thoikhoabieu',
            'verbose_name_plural': 'Thời khóa biểu',
        },
        'segment': ['xem-tkb'],
    }
    
    if not ma_dot:
        return render(request, 'admin/thoikhoabieu.html', context)
    
    try:
        dot_xep = DotXep.objects.get(ma_dot=ma_dot)
        context['dot_xep'] = dot_xep
        
        if view_type == 'teacher':
            # Lấy danh sách giáo viên CÓ LỊCH DẠY THỰC TẾ trong đợt này
            # Lấy từ TKB thông qua PhanCong
            gv_query = PhanCong.objects.filter(
                ma_dot=dot_xep,
                ma_gv__isnull=False,
                ma_lop__tkb_list__ma_dot=dot_xep,
                ma_lop__tkb_list__is_deleted=False
            ).select_related('ma_gv', 'ma_gv__ma_bo_mon', 'ma_gv__ma_bo_mon__ma_khoa')
            
            # Filter theo khoa nếu có
            if ma_khoa:
                gv_query = gv_query.filter(ma_gv__ma_bo_mon__ma_khoa__ma_khoa=ma_khoa)
            
            # Filter theo bộ môn nếu có
            if ma_bo_mon:
                gv_query = gv_query.filter(ma_gv__ma_bo_mon__ma_bo_mon=ma_bo_mon)
            
            gv_co_lich = gv_query.distinct().values(
                'ma_gv__ma_gv', 'ma_gv__ten_gv', 'ma_gv__ma_bo_mon__ma_khoa__ten_khoa'
            ).order_by('ma_gv__ma_gv')
            
            teachers = [{
                'ma_gv': gv['ma_gv__ma_gv'], 
                'ten_gv': gv['ma_gv__ten_gv'],
                'ten_khoa': gv['ma_gv__ma_bo_mon__ma_khoa__ten_khoa']
            } for gv in gv_co_lich]
            context['teachers'] = teachers
            
            # Không tự động chọn GV, bắt buộc phải chọn thủ công
            # Nếu chưa chọn GV, không hiển thị lịch
            if selected_id:
                context['selected_id'] = selected_id
            
            if selected_id:
                try:
                    gv = GiangVien.objects.get(ma_gv=selected_id)
                    context['selected_teacher'] = gv
                    
                    # Lấy các lớp mà GV dạy trong đợt này
                    lop_gv = PhanCong.objects.filter(
                        ma_dot=dot_xep, ma_gv=gv
                    ).values_list('ma_lop__ma_lop', flat=True)
                    
                    # Lấy TKB của các lớp đó
                    tkb_list = ThoiKhoaBieu.objects.filter(
                        ma_dot=dot_xep,
                        ma_lop__ma_lop__in=lop_gv,
                        is_deleted=False
                    ).select_related(
                        'ma_lop', 'ma_lop__ma_mon_hoc', 'ma_phong', 
                        'time_slot_id', 'time_slot_id__ca'
                    ).order_by('time_slot_id__thu', 'time_slot_id__ca')
                    
                    schedule_data = build_schedule_data(
                        tkb_list, display_mode, week_number, dot_xep
                    )
                    context['schedule_data'] = schedule_data
                    
                    # Thêm ngày tháng cho từng thứ nếu ở chế độ weekly
                    if display_mode == 'weekly':
                        context['week_dates'] = get_week_dates(dot_xep, week_number)
                except GiangVien.DoesNotExist:
                    context['error'] = f'Không tìm thấy giáo viên {selected_id}'
                
        else:  # view_type == 'room'
            # Lấy danh sách phòng CÓ LỊCH SỬ DỤNG trong đợt này
            room_query = PhongHoc.objects.filter(
                tkb_list__ma_dot=dot_xep,
                tkb_list__is_deleted=False
            )
            
            # Filter theo khoa nếu có (qua PhanCong của lớp)
            if ma_khoa:
                room_query = room_query.filter(
                    tkb_list__ma_lop__phan_cong_list__ma_gv__ma_bo_mon__ma_khoa__ma_khoa=ma_khoa,
                    tkb_list__ma_lop__phan_cong_list__ma_dot=dot_xep
                )
            
            # Filter theo bộ môn nếu có
            if ma_bo_mon:
                room_query = room_query.filter(
                    tkb_list__ma_lop__phan_cong_list__ma_gv__ma_bo_mon__ma_bo_mon=ma_bo_mon,
                    tkb_list__ma_lop__phan_cong_list__ma_dot=dot_xep
                )
            
            rooms = room_query.distinct().order_by('ma_phong')
            context['rooms'] = rooms
            
            # Không tự động chọn phòng, bắt buộc phải chọn thủ công
            # Nếu chưa chọn phòng, không hiển thị lịch
            if selected_id:
                context['selected_id'] = selected_id
            
            if selected_id:
                try:
                    phong = PhongHoc.objects.get(ma_phong=selected_id)
                    context['selected_room'] = phong
                    
                    # Lấy TKB của phòng
                    tkb_list = ThoiKhoaBieu.objects.filter(
                        ma_dot=dot_xep,
                        ma_phong=phong,
                        is_deleted=False
                    ).select_related(
                        'ma_lop', 'ma_lop__ma_mon_hoc', 'ma_phong',
                        'time_slot_id', 'time_slot_id__ca'
                    ).order_by('time_slot_id__thu', 'time_slot_id__ca')
                    
                    schedule_data = build_schedule_data(
                        tkb_list, display_mode, week_number, dot_xep
                    )
                    context['schedule_data'] = schedule_data
                    
                    # Thêm ngày tháng cho từng thứ nếu ở chế độ weekly
                    if display_mode == 'weekly':
                        context['week_dates'] = get_week_dates(dot_xep, week_number)
                except PhongHoc.DoesNotExist:
                    context['error'] = f'Không tìm thấy phòng {selected_id}'
                except PhongHoc.DoesNotExist:
                    context['error'] = f'Không tìm thấy phòng {selected_id}'
        
    except DotXep.DoesNotExist:
        context['error'] = f'Không tìm thấy đợt xếp lịch {ma_dot}'
    except Exception as e:
        logger.exception(f"Lỗi khi hiển thị TKB: {e}")
        context['error'] = f'Lỗi: {str(e)}'
    
    return render(request, 'admin/thoikhoabieu.html', context)


def build_schedule_data(tkb_list, display_mode, week_number, dot_xep):
    """
    Xây dựng dữ liệu lịch học theo tuần
    Args:
        tkb_list: QuerySet các TKB
        display_mode: 'general' hoặc 'weekly'
        week_number: Số tuần hiện tại
        dot_xep: DotXep object
    Returns: {
        'schedule': {
            'thu_2': [{'ca': 1, 'ca_info': {...}, 'classes': [...]}, ...],
            'thu_3': [...],
            ...
        },
        'ca_list': [...]
    }
    """
    # Khởi tạo cấu trúc dữ liệu
    schedule = {f'thu_{i}': {} for i in range(2, 9)}  # Thứ 2-8 (8=CN)
    
    # Lấy danh sách ca học
    ca_list = KhungTG.objects.all().order_by('ma_khung_gio')
    
    # Khởi tạo tất cả các slot trống
    for thu in range(2, 9):
        thu_key = f'thu_{thu}'
        for ca in ca_list:
            schedule[thu_key][ca.ma_khung_gio] = {
                'ca': ca.ma_khung_gio,
                'ca_info': {
                    'ten_ca': ca.ten_ca,
                    'gio_bd': ca.gio_bat_dau.strftime('%H:%M'),
                    'gio_kt': ca.gio_ket_thuc.strftime('%H:%M'),
                },
                'classes': []
            }
    
    # Tạo cache cho PhanCong để tránh query nhiều lần
    phan_cong_cache = {}
    phan_cong_data = PhanCong.objects.filter(
        ma_dot=dot_xep
    ).select_related('ma_gv', 'ma_lop')
    
    for pc in phan_cong_data:
        phan_cong_cache[pc.ma_lop.ma_lop] = {
            'gv_name': pc.ma_gv.ten_gv if pc.ma_gv else 'Chưa phân',
            'gv_code': pc.ma_gv.ma_gv if pc.ma_gv else '',
        }
    
    # Điền dữ liệu từ TKB
    for tkb in tkb_list:
        thu = tkb.time_slot_id.thu
        ca = tkb.time_slot_id.ca.ma_khung_gio
        thu_key = f'thu_{thu}'
        
        # Parse tuần học
        weeks = parse_tuan_hoc(tkb.tuan_hoc, week_number, display_mode)
        
        # Nếu ở chế độ chi tiết theo tuần và không có buổi nào trong tuần này thì bỏ qua
        if display_mode == 'weekly' and week_number not in weeks:
            continue
        
        # Lấy thông tin giáo viên từ cache
        gv_info = phan_cong_cache.get(tkb.ma_lop.ma_lop, {
            'gv_name': 'N/A',
            'gv_code': ''
        })
        
        class_info = {
            'ma_tkb': tkb.ma_tkb,  # Thêm mã TKB để có thể edit/delete
            'ma_lop': tkb.ma_lop.ma_lop,
            'mon_hoc': tkb.ma_lop.ma_mon_hoc.ten_mon_hoc,
            'ma_mon': tkb.ma_lop.ma_mon_hoc.ma_mon_hoc,
            'phong': tkb.ma_phong.ma_phong if tkb.ma_phong else 'TBA',
            'gv_name': gv_info['gv_name'],
            'gv_code': gv_info['gv_code'],
            'weeks': weeks,
            'week_display': format_weeks(weeks) if display_mode == 'general' else f'Tuần {week_number}',
            'tuan_hoc': tkb.tuan_hoc or ('1' * 15),  # Pattern tuần học để edit
        }
        
        schedule[thu_key][ca]['classes'].append(class_info)
    
    # Chuyển dict thành list để dễ iterate trong template
    result_schedule = {}
    for thu_key, ca_dict in schedule.items():
        result_schedule[thu_key] = [slot_data for ca_id, slot_data in sorted(ca_dict.items())]
    
    return {
        'schedule': result_schedule,
        'ca_list': list(ca_list.values('ma_khung_gio', 'ten_ca', 'gio_bat_dau', 'gio_ket_thuc'))
    }


def parse_tuan_hoc(tuan_hoc_pattern, week_number, display_mode):
    """
    Parse chuỗi pattern tuần học (VD: "1111111000000000") thành list các tuần
    Returns: [1, 2, 3, 4, 5, 6, 7] cho pattern trên
    """
    if not tuan_hoc_pattern:
        # Mặc định: tất cả 15 tuần
        return list(range(1, 16))
    
    weeks = []
    for i, char in enumerate(tuan_hoc_pattern):
        if char == '1':
            weeks.append(i + 1)
    
    return weeks if weeks else list(range(1, 16))


def format_weeks(weeks):
    """
    Format danh sách tuần thành chuỗi ngắn gọn
    VD: [1,2,3,4,5,7,8] -> "T1-5, 7-8"
    """
    if not weeks:
        return ""
    
    weeks = sorted(weeks)
    ranges = []
    start = weeks[0]
    end = weeks[0]
    
    for i in range(1, len(weeks)):
        if weeks[i] == end + 1:
            end = weeks[i]
        else:
            if start == end:
                ranges.append(f"T{start}")
            else:
                ranges.append(f"T{start}-{end}")
            start = weeks[i]
            end = weeks[i]
    
    # Thêm range cuối cùng
    if start == end:
        ranges.append(f"T{start}")
    else:
        ranges.append(f"T{start}-{end}")
    
    return ", ".join(ranges)


def get_week_dates(dot_xep, week_number):
    """
    Tính ngày cụ thể cho từng thứ trong tuần
    Returns: {
        2: {'date': datetime, 'display': '01/01'},
        3: {'date': datetime, 'display': '02/01'},
        ...
        8: {'date': datetime, 'display': '07/01'}
    }
    """
    # Lấy ngày bắt đầu từ DuKienDT
    if not dot_xep.ma_du_kien_dt or not dot_xep.ma_du_kien_dt.ngay_bd:
        return {}
    
    # Tính ngày bắt đầu của tuần (Thứ 2)
    # week_number = 1 => tuần đầu tiên
    start_date = dot_xep.ma_du_kien_dt.ngay_bd
    days_to_add = (week_number - 1) * 7
    week_start = start_date + timedelta(days=days_to_add)
    
    # Điều chỉnh để week_start là thứ 2
    # weekday(): 0=Monday, 6=Sunday
    weekday = week_start.weekday()
    if weekday != 0:  # Nếu không phải thứ 2
        week_start = week_start - timedelta(days=weekday)
    
    week_dates = {}
    for thu in range(2, 9):  # Thứ 2-8 (8=CN)
        if thu == 8:
            # Chủ nhật
            day_offset = 6
        else:
            # Thứ 2-7
            day_offset = thu - 2
        
        day_date = week_start + timedelta(days=day_offset)
        week_dates[thu] = {
            'date': day_date,
            'display': day_date.strftime('%d/%m')
        }
    
    return week_dates


# ==================== VALIDATION FUNCTIONS ====================

def validate_tkb_constraints(ma_dot, ma_lop, ma_phong, time_slot_id, ma_gv=None, exclude_ma_tkb=None):
    """
    Kiểm tra các ràng buộc khi thêm/sửa TKB
    Returns: {
        'valid': True/False,
        'errors': [],  # Ràng buộc cứng bị vi phạm
        'warnings': []  # Ràng buộc mềm bị vi phạm
    }
    """
    errors = []
    warnings = []
    
    try:
        dot_xep = DotXep.objects.get(ma_dot=ma_dot)
        lop_mon_hoc = LopMonHoc.objects.get(ma_lop=ma_lop)
        phong_hoc = PhongHoc.objects.get(ma_phong=ma_phong) if ma_phong else None
        time_slot = TimeSlot.objects.get(time_slot_id=time_slot_id)
        
        # Lấy GV từ phân công nếu không truyền vào
        if not ma_gv:
            phan_cong = PhanCong.objects.filter(
                ma_dot=dot_xep, ma_lop=lop_mon_hoc
            ).first()
            if phan_cong and phan_cong.ma_gv:
                ma_gv = phan_cong.ma_gv.ma_gv
        
        giang_vien = GiangVien.objects.get(ma_gv=ma_gv) if ma_gv else None
        
        # 1. RÀNG BUỘC CỨNG: Kiểm tra trùng phòng cùng thời gian
        if phong_hoc:
            conflicting_room = ThoiKhoaBieu.objects.filter(
                ma_dot=dot_xep,
                ma_phong=phong_hoc,
                time_slot_id=time_slot,
                is_deleted=False
            )
            if exclude_ma_tkb:
                conflicting_room = conflicting_room.exclude(ma_tkb=exclude_ma_tkb)
            
            if conflicting_room.exists():
                conflict = conflicting_room.first()
                errors.append(
                    f"❌ Phòng {phong_hoc.ma_phong} đã bị trùng với lớp "
                    f"{conflict.ma_lop.ma_lop} ({conflict.ma_lop.ma_mon_hoc.ten_mon_hoc}) "
                    f"vào {time_slot}"
                )
        
        # 2. RÀNG BUỘC CỨNG: Kiểm tra trùng giáo viên cùng thời gian
        if giang_vien:
            # Lấy tất cả các lớp GV dạy trong đợt này
            lop_cua_gv = PhanCong.objects.filter(
                ma_dot=dot_xep, ma_gv=giang_vien
            ).values_list('ma_lop__ma_lop', flat=True)
            
            conflicting_teacher = ThoiKhoaBieu.objects.filter(
                ma_dot=dot_xep,
                ma_lop__ma_lop__in=lop_cua_gv,
                time_slot_id=time_slot,
                is_deleted=False
            )
            if exclude_ma_tkb:
                conflicting_teacher = conflicting_teacher.exclude(ma_tkb=exclude_ma_tkb)
            
            if conflicting_teacher.exists():
                conflict = conflicting_teacher.first()
                errors.append(
                    f"❌ GV {giang_vien.ten_gv} đã có lịch dạy lớp "
                    f"{conflict.ma_lop.ma_lop} ({conflict.ma_lop.ma_mon_hoc.ten_mon_hoc}) "
                    f"vào {time_slot}"
                )
        
        # 3. RÀNG BUỘC CỨNG: Kiểm tra phòng phù hợp với loại môn (LT/TH)
        if phong_hoc and lop_mon_hoc:
            mon_hoc = lop_mon_hoc.ma_mon_hoc
            # Kiểm tra nếu môn có thực hành nhưng phòng không phải phòng TH
            if mon_hoc.so_tiet_th and mon_hoc.so_tiet_th > 0:
                if phong_hoc.loai_phong and 'TH' not in phong_hoc.loai_phong.upper() and 'MÁY' not in phong_hoc.loai_phong.upper():
                    warnings.append(
                        f"⚠️ Môn {mon_hoc.ten_mon_hoc} có {mon_hoc.so_tiet_th} tiết TH "
                        f"nhưng phòng {phong_hoc.ma_phong} là {phong_hoc.loai_phong or 'không xác định'}"
                    )
            
            # Kiểm tra nếu chỉ có lý thuyết nhưng lại dùng phòng máy
            if (not mon_hoc.so_tiet_th or mon_hoc.so_tiet_th == 0) and mon_hoc.so_tiet_lt:
                if phong_hoc.loai_phong and ('TH' in phong_hoc.loai_phong.upper() or 'MÁY' in phong_hoc.loai_phong.upper()):
                    warnings.append(
                        f"⚠️ Môn {mon_hoc.ten_mon_hoc} chỉ có lý thuyết "
                        f"nhưng đang xếp vào phòng {phong_hoc.ma_phong} ({phong_hoc.loai_phong})"
                    )
        
        # 4. RÀNG BUỘC CỨNG: Kiểm tra sức chứa phòng
        if phong_hoc and phong_hoc.suc_chua and lop_mon_hoc.so_luong_sv:
            if lop_mon_hoc.so_luong_sv > phong_hoc.suc_chua:
                errors.append(
                    f"❌ Lớp có {lop_mon_hoc.so_luong_sv} SV nhưng phòng "
                    f"{phong_hoc.ma_phong} chỉ chứa được {phong_hoc.suc_chua} người"
                )
        
        # 5. RÀNG BUỘC MỀM: Kiểm tra giờ làm việc của GV (số tiết/tuần)
        if giang_vien:
            # Đếm số tiết GV đã dạy trong tuần
            total_slots = ThoiKhoaBieu.objects.filter(
                ma_dot=dot_xep,
                ma_lop__ma_lop__in=lop_cua_gv,
                is_deleted=False
            ).count()
            
            if exclude_ma_tkb:
                # Nếu đang sửa, không tính slot hiện tại
                pass
            else:
                # Nếu đang thêm mới
                total_slots += 1
            
            # Giả sử mỗi slot = 1 ca = 3 tiết, tối đa 10 ca/tuần = 30 tiết
            MAX_SLOTS_PER_WEEK = 10
            if total_slots > MAX_SLOTS_PER_WEEK:
                warnings.append(
                    f"⚠️ GV {giang_vien.ten_gv} đã có {total_slots - 1} ca dạy, "
                    f"nếu thêm ca này sẽ là {total_slots} ca (khuyến nghị tối đa {MAX_SLOTS_PER_WEEK} ca/tuần)"
                )
        
        # 6. RÀNG BUỘC MỀM: Kiểm tra nguyện vọng của GV
        if giang_vien:
            nguyen_vong = NguyenVong.objects.filter(
                ma_dot=dot_xep,
                ma_gv=giang_vien,
                time_slot_id=time_slot
            ).exists()
            
            if nguyen_vong:
                # GV đã đăng ký nguyện vọng muốn dạy slot này
                pass
            else:
                # GV chưa đăng ký nguyện vọng cho slot này
                warnings.append(
                    f"ℹ️ GV {giang_vien.ten_gv} chưa đăng ký nguyện vọng cho {time_slot}"
                )
        
        # 7. Kiểm tra xem GV có đủ điều kiện dạy môn không
        if giang_vien and lop_mon_hoc:
            can_teach = GVDayMon.objects.filter(
                ma_gv=giang_vien,
                ma_mon_hoc=lop_mon_hoc.ma_mon_hoc
            ).exists()
            
            if not can_teach:
                warnings.append(
                    f"⚠️ GV {giang_vien.ten_gv} chưa được đăng ký là người có thể dạy "
                    f"môn {lop_mon_hoc.ma_mon_hoc.ten_mon_hoc}"
                )
        
    except (DotXep.DoesNotExist, LopMonHoc.DoesNotExist, PhongHoc.DoesNotExist, 
            TimeSlot.DoesNotExist, GiangVien.DoesNotExist) as e:
        errors.append(f"❌ Lỗi dữ liệu: {str(e)}")
    
    return {
        'valid': len(errors) == 0,
        'errors': errors,
        'warnings': warnings
    }


# ==================== API ENDPOINTS FOR CRUD ====================

@csrf_exempt
@require_http_methods(["POST"])
def tkb_create_api(request):
    """API tạo mới một bản ghi TKB - tự động tạo lớp môn học mới"""
    try:
        data = json.loads(request.body)
        ma_dot = data.get('ma_dot')
        ma_mon_hoc = data.get('ma_mon_hoc')  # Đổi từ ma_lop sang ma_mon_hoc
        ma_gv = data.get('ma_gv')  # Mã giáo viên
        nhom_mh = int(data.get('nhom_mh', 1))
        to_mh = int(data.get('to_mh', 0))
        so_luong_sv = int(data.get('so_luong_sv', 40))
        ma_phong = data.get('ma_phong')
        time_slot_id = data.get('time_slot_id')
        tuan_hoc = data.get('tuan_hoc', '1' * 15)  # Mặc định tất cả 15 tuần
        
        if not all([ma_dot, ma_mon_hoc, ma_gv, ma_phong, time_slot_id]):
            return JsonResponse({
                'status': 'error',
                'message': 'Thiếu thông tin bắt buộc'
            }, status=400)
        
        # Lấy objects
        dot_xep = DotXep.objects.get(ma_dot=ma_dot)
        mon_hoc = MonHoc.objects.get(ma_mon_hoc=ma_mon_hoc)
        giang_vien = GiangVien.objects.get(ma_gv=ma_gv)
        ts = TimeSlot.objects.get(time_slot_id=time_slot_id)
        
        # Tạo mã lớp: MAHOC_Nhom_To (VD: TOAN_1_0, CNTT_2_1)
        if to_mh > 0:
            ma_lop = f"{ma_mon_hoc}_{nhom_mh}_{to_mh}"
        else:
            ma_lop = f"{ma_mon_hoc}_{nhom_mh}"
        
        # Kiểm tra lớp đã tồn tại chưa
        lop, created = LopMonHoc.objects.get_or_create(
            ma_lop=ma_lop,
            defaults={
                'ma_mon_hoc': mon_hoc,
                'nhom_mh': nhom_mh,
                'to_mh': to_mh,
                'so_luong_sv': so_luong_sv,
                'so_ca_tuan': 1,
            }
        )
        
        if not created:
            # Lớp đã tồn tại, cập nhật thông tin
            lop.so_luong_sv = so_luong_sv
            lop.save()
        
        # Tạo hoặc cập nhật phân công
        phan_cong, pc_created = PhanCong.objects.get_or_create(
            ma_dot=dot_xep,
            ma_lop=lop,
            defaults={
                'ma_gv': giang_vien,
            }
        )
        
        if not pc_created and phan_cong.ma_gv != giang_vien:
            phan_cong.ma_gv = giang_vien
            phan_cong.save()
        
        # Validate ràng buộc
        validation = validate_tkb_constraints(
            ma_dot, lop.ma_lop, ma_phong, time_slot_id
        )
        
        if not validation['valid']:
            return JsonResponse({
                'status': 'error',
                'message': 'Vi phạm ràng buộc cứng',
                'errors': validation['errors'],
                'warnings': validation['warnings']
            }, status=400)
        
        # Tạo mã TKB siêu ngắn (max 15 ký tự)
        # Vì cột MaTKB chỉ có max_length=15 trong database
        # Format: {thu}{ca}{hash6}
        # VD: 21A3F5B8 = Thứ 2, Ca 1, Hash 6 ký tự
        
        import hashlib
        import re
        
        # Lấy thứ và ca từ timeslot (VD: Thu2-Ca1 → 2, 1)
        ts_match = re.match(r'Thu(\d+)-Ca(\d+)', time_slot_id)
        if ts_match:
            thu = ts_match.group(1)
            ca = ts_match.group(2)
        else:
            # Fallback nếu format khác (CN-Ca1 → 8, 1)
            thu = '8' if 'CN' in time_slot_id else '0'
            ca_match = re.search(r'Ca(\d+)', time_slot_id)
            ca = ca_match.group(1) if ca_match else '0'
        
        # Hash từ đợt + lớp để đảm bảo unique
        hash_input = f"{ma_dot}_{ma_lop}_{time_slot_id}".encode()
        hash_hex = hashlib.md5(hash_input).hexdigest()[:6].upper()
        
        # Mã TKB: {thu}{ca}{hash} (VD: 21A3F5B8 = 8 ký tự)
        ma_tkb = f"{thu}{ca}{hash_hex}"
        
        # Kiểm tra TKB đã tồn tại chưa
        if ThoiKhoaBieu.objects.filter(ma_tkb=ma_tkb, is_deleted=False).exists():
            return JsonResponse({
                'status': 'error',
                'message': f'Lịch này đã tồn tại: {ma_lop} - {time_slot_id}'
            }, status=400)
        
        # Lấy ngày bắt đầu/kết thúc từ đợt
        ngay_bd = dot_xep.ma_du_kien_dt.ngay_bd if dot_xep.ma_du_kien_dt else None
        ngay_kt = dot_xep.ma_du_kien_dt.ngay_kt if dot_xep.ma_du_kien_dt else None
        
        # Tạo TKB
        tkb = ThoiKhoaBieu.objects.create(
            ma_tkb=ma_tkb,
            ma_dot=dot_xep,
            ma_lop=lop,
            ma_phong_id=ma_phong,
            time_slot_id=ts,
            tuan_hoc=tuan_hoc,
            ngay_bd=ngay_bd,
            ngay_kt=ngay_kt,
            is_deleted=False
        )
        
        # Log
        new_data = {
            'ma_tkb': tkb.ma_tkb,
            'ma_lop': tkb.ma_lop.ma_lop,
            'ma_mon_hoc': mon_hoc.ma_mon_hoc,
            'ma_gv': giang_vien.ma_gv,
            'ma_phong': tkb.ma_phong.ma_phong if tkb.ma_phong else None,
            'time_slot_id': tkb.time_slot_id.time_slot_id,
            'tuan_hoc': tkb.tuan_hoc,
        }
        
        TKBLog.objects.create(
            ma_tkb=ma_tkb,
            action='CREATE',
            user=request.user.username if request.user.is_authenticated else 'anonymous',
            old_data=None,
            new_data=new_data,
            reason='Tạo lịch mới'
        )
        
        return JsonResponse({
            'status': 'success',
            'message': f'Thêm lịch thành công! Lớp: {ma_lop}, GV: {giang_vien.ten_gv}',
            'warnings': validation.get('warnings', []),
            'data': {
                'ma_tkb': tkb.ma_tkb,
                'ma_lop': tkb.ma_lop.ma_lop,
                'ma_mon_hoc': mon_hoc.ten_mon_hoc,
                'ma_gv': giang_vien.ten_gv,
                'ma_phong': tkb.ma_phong.ma_phong if tkb.ma_phong else None,
                'time_slot': str(tkb.time_slot_id),
                'created_new_class': created
            }
        })
        
    except Exception as e:
        logger.exception(f"Lỗi khi tạo TKB: {e}")
        return JsonResponse({
            'status': 'error',
            'message': f'Lỗi: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def tkb_update_api(request):
    """API cập nhật TKB - cho phép thay đổi môn học, phòng, timeslot"""
    try:
        data = json.loads(request.body)
        ma_tkb = data.get('ma_tkb')
        ma_mon_hoc = data.get('ma_mon_hoc')
        nhom = data.get('nhom')
        to = data.get('to')
        so_luong_sv = data.get('so_luong_sv')
        ma_phong = data.get('ma_phong')
        time_slot_id = data.get('time_slot_id')
        tuan_hoc = data.get('tuan_hoc', '')
        
        tkb = ThoiKhoaBieu.objects.get(ma_tkb=ma_tkb, is_deleted=False)
        old_ma_lop = tkb.ma_lop.ma_lop
        
        # Nếu thay đổi môn học → tạo mã lớp mới
        new_ma_lop = old_ma_lop
        if ma_mon_hoc and nhom and to is not None:
            new_ma_lop = f"{ma_mon_hoc}_{nhom}_{to}"
            
            # Kiểm tra xem lớp mới đã tồn tại chưa (ngoại trừ lớp hiện tại)
            if new_ma_lop != old_ma_lop:
                existing_lop = LopMonHoc.objects.filter(ma_lop=new_ma_lop).first()
                if not existing_lop:
                    # Tạo lớp mới
                    mon_hoc = MonHoc.objects.get(ma_mon_hoc=ma_mon_hoc)
                    existing_lop = LopMonHoc.objects.create(
                        ma_lop=new_ma_lop,
                        ma_mon_hoc=mon_hoc,
                        nhom=int(nhom),
                        to=int(to),
                        so_luong_sv=int(so_luong_sv) if so_luong_sv else None
                    )
                    logger.info(f"Tạo lớp mới khi edit: {new_ma_lop}")
                
                # Cập nhật lớp
                tkb.ma_lop = existing_lop
        
        # Cập nhật số lượng SV nếu có
        if so_luong_sv and tkb.ma_lop:
            tkb.ma_lop.so_luong_sv = int(so_luong_sv)
            tkb.ma_lop.save()
        
        # Validate với exclude current
        validation = validate_tkb_constraints(
            tkb.ma_dot.ma_dot,
            tkb.ma_lop.ma_lop,
            ma_phong if ma_phong else (tkb.ma_phong.ma_phong if tkb.ma_phong else None),
            time_slot_id if time_slot_id else tkb.time_slot_id.time_slot_id,
            exclude_ma_tkb=ma_tkb
        )
        
        if not validation['valid']:
            return JsonResponse({
                'status': 'error',
                'message': 'Vi phạm ràng buộc cứng',
                'errors': validation['errors'],
                'warnings': validation['warnings']
            }, status=400)
        
        # Lưu old data
        old_data = {
            'ma_lop': old_ma_lop,
            'ma_phong': tkb.ma_phong.ma_phong if tkb.ma_phong else None,
            'time_slot_id': tkb.time_slot_id.time_slot_id,
            'tuan_hoc': tkb.tuan_hoc,
        }
        
        # Update phòng và timeslot
        if ma_phong:
            tkb.ma_phong_id = ma_phong
        if time_slot_id:
            tkb.time_slot_id_id = time_slot_id
        if tuan_hoc is not None:
            tkb.tuan_hoc = tuan_hoc
        
        tkb.save()
        
        # Log
        new_data = {
            'ma_lop': tkb.ma_lop.ma_lop,
            'ma_phong': tkb.ma_phong.ma_phong if tkb.ma_phong else None,
            'time_slot_id': tkb.time_slot_id.time_slot_id,
            'tuan_hoc': tkb.tuan_hoc,
        }
        
        change_summary = []
        if old_ma_lop != tkb.ma_lop.ma_lop:
            change_summary.append(f"Lớp: {old_ma_lop} → {tkb.ma_lop.ma_lop}")
        if old_data['ma_phong'] != new_data['ma_phong']:
            change_summary.append(f"Phòng: {old_data['ma_phong']} → {new_data['ma_phong']}")
        if old_data['time_slot_id'] != new_data['time_slot_id']:
            change_summary.append(f"Timeslot: {old_data['time_slot_id']} → {new_data['time_slot_id']}")
        
        TKBLog.objects.create(
            ma_tkb=ma_tkb,
            action='UPDATE',
            user=request.user.username if request.user.is_authenticated else 'anonymous',
            old_data=old_data,
            new_data=new_data,
            reason='Cập nhật: ' + ', '.join(change_summary)
        )
        
        message = 'Cập nhật lịch thành công'
        if change_summary:
            message += ': ' + ', '.join(change_summary)
        
        return JsonResponse({
            'status': 'success',
            'message': message,
            'warnings': validation['warnings']
        })
        
    except ThoiKhoaBieu.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Không tìm thấy lịch học'
        }, status=404)
    except MonHoc.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Không tìm thấy môn học'
        }, status=404)
    except Exception as e:
        logger.exception(f"Lỗi khi cập nhật TKB: {e}")
        return JsonResponse({
            'status': 'error',
            'message': f'Lỗi: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def tkb_delete_api(request):
    """API xóa TKB (soft delete)"""
    try:
        data = json.loads(request.body)
        ma_tkb = data.get('ma_tkb')
        reason = data.get('reason', '')
        
        tkb = ThoiKhoaBieu.objects.get(ma_tkb=ma_tkb)
        
        # Lưu dữ liệu cũ trước khi xóa
        old_data = {
            'ma_tkb': tkb.ma_tkb,
            'ma_lop': tkb.ma_lop.ma_lop,
            'ma_phong': tkb.ma_phong.ma_phong if tkb.ma_phong else None,
            'time_slot_id': tkb.time_slot_id.time_slot_id,
            'tuan_hoc': tkb.tuan_hoc,
        }
        
        # Soft delete
        tkb.is_deleted = True
        tkb.save()
        
        # Log
        TKBLog.objects.create(
            ma_tkb=ma_tkb,
            action='DELETE',
            user=request.user.username if request.user.is_authenticated else 'anonymous',
            old_data=old_data,
            new_data={'is_deleted': True},
            reason=reason
        )
        
        return JsonResponse({
            'status': 'success',
            'message': 'Xóa lịch thành công (có thể phục hồi)'
        })
        
    except ThoiKhoaBieu.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Không tìm thấy lịch học'
        }, status=404)
    except Exception as e:
        logger.exception(f"Lỗi khi xóa TKB: {e}")
        return JsonResponse({
            'status': 'error',
            'message': f'Lỗi: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def tkb_restore_api(request):
    """API phục hồi TKB đã xóa"""
    try:
        data = json.loads(request.body)
        ma_tkb = data.get('ma_tkb')
        
        tkb = ThoiKhoaBieu.objects.get(ma_tkb=ma_tkb)
        
        if not tkb.is_deleted:
            return JsonResponse({
                'status': 'error',
                'message': 'Lịch này chưa bị xóa'
            }, status=400)
        
        # Lấy GV từ phân công
        phan_cong = PhanCong.objects.filter(
            ma_dot=tkb.ma_dot,
            ma_lop=tkb.ma_lop
        ).first()
        ma_gv = phan_cong.ma_gv.ma_gv if phan_cong and phan_cong.ma_gv else None
        
        # Validate xem slot có bị trùng không (cả phòng và GV)
        validation = validate_tkb_constraints(
            tkb.ma_dot.ma_dot,
            tkb.ma_lop.ma_lop,
            tkb.ma_phong.ma_phong if tkb.ma_phong else None,
            tkb.time_slot_id.time_slot_id,
            ma_gv=ma_gv,
            exclude_ma_tkb=ma_tkb
        )
        
        if not validation['valid']:
            return JsonResponse({
                'status': 'error',
                'message': 'Không thể phục hồi vì vi phạm ràng buộc',
                'errors': validation['errors']
            }, status=400)
        
        # Phục hồi
        tkb.is_deleted = False
        tkb.save()
        
        # Log
        TKBLog.objects.create(
            ma_tkb=ma_tkb,
            action='RESTORE',
            user=request.user.username if request.user.is_authenticated else 'anonymous',
            old_data={'is_deleted': True},
            new_data={'is_deleted': False},
            reason='Phục hồi lịch đã xóa'
        )
        
        return JsonResponse({
            'status': 'success',
            'message': 'Phục hồi lịch thành công',
            'warnings': validation['warnings']
        })
        
    except ThoiKhoaBieu.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Không tìm thấy lịch học'
        }, status=404)
    except Exception as e:
        logger.exception(f"Lỗi khi phục hồi TKB: {e}")
        return JsonResponse({
            'status': 'error',
            'message': f'Lỗi: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def tkb_swap_api(request):
    """API hoán đổi 2 TKB với tùy chọn swap phòng"""
    try:
        data = json.loads(request.body)
        ma_tkb_1 = data.get('ma_tkb_1')
        ma_tkb_2 = data.get('ma_tkb_2')
        swap_phong = data.get('swap_phong', True)  # Mặc định có swap phòng
        
        tkb1 = ThoiKhoaBieu.objects.get(ma_tkb=ma_tkb_1, is_deleted=False)
        tkb2 = ThoiKhoaBieu.objects.get(ma_tkb=ma_tkb_2, is_deleted=False)
        
        # Kiểm tra xem 2 TKB có cùng GV không
        gv1 = PhanCong.objects.filter(ma_dot=tkb1.ma_dot, ma_lop=tkb1.ma_lop).first()
        gv2 = PhanCong.objects.filter(ma_dot=tkb2.ma_dot, ma_lop=tkb2.ma_lop).first()
        
        same_teacher = gv1 and gv2 and gv1.ma_gv == gv2.ma_gv
        
        # Lưu data cũ
        old_data_1 = {
            'ma_phong': tkb1.ma_phong.ma_phong if tkb1.ma_phong else None,
            'time_slot_id': tkb1.time_slot_id.time_slot_id,
            'ma_lop': tkb1.ma_lop.ma_lop,
            'gv': gv1.ma_gv.ma_gv if gv1 and gv1.ma_gv else None
        }
        old_data_2 = {
            'ma_phong': tkb2.ma_phong.ma_phong if tkb2.ma_phong else None,
            'time_slot_id': tkb2.time_slot_id.time_slot_id,
            'ma_lop': tkb2.ma_lop.ma_lop,
            'gv': gv2.ma_gv.ma_gv if gv2 and gv2.ma_gv else None
        }
        
        warnings = []
        errors = []
        
        # Hoán đổi timeslot (luôn luôn)
        temp_slot = tkb1.time_slot_id
        tkb1.time_slot_id = tkb2.time_slot_id
        tkb2.time_slot_id = temp_slot
        
        # Hoán đổi phòng (tùy chọn)
        if swap_phong:
            temp_phong = tkb1.ma_phong
            tkb1.ma_phong = tkb2.ma_phong
            tkb2.ma_phong = temp_phong
            
            # Validate phòng mới
            # 1. Kiểm tra phòng trùng timeslot
            if tkb1.ma_phong:
                conflict1 = ThoiKhoaBieu.objects.filter(
                    ma_dot=tkb1.ma_dot,
                    ma_phong=tkb1.ma_phong,
                    time_slot_id=tkb1.time_slot_id,
                    is_deleted=False
                ).exclude(ma_tkb=ma_tkb_1)
                
                if conflict1.exists():
                    errors.append(f"❌ Phòng {tkb1.ma_phong.ma_phong} đã bị trùng tại {tkb1.time_slot_id}")
            
            if tkb2.ma_phong:
                conflict2 = ThoiKhoaBieu.objects.filter(
                    ma_dot=tkb2.ma_dot,
                    ma_phong=tkb2.ma_phong,
                    time_slot_id=tkb2.time_slot_id,
                    is_deleted=False
                ).exclude(ma_tkb=ma_tkb_2)
                
                if conflict2.exists():
                    errors.append(f"❌ Phòng {tkb2.ma_phong.ma_phong} đã bị trùng tại {tkb2.time_slot_id}")
            
            # 2. Kiểm tra loại phòng phù hợp (chỉ warning)
            if tkb1.ma_phong:
                mon1 = tkb1.ma_lop.ma_mon_hoc
                loai_phong1 = tkb1.ma_phong.loai_phong or ''
                
                if mon1.so_tiet_th and mon1.so_tiet_th > 0:
                    # Môn TH nên dùng phòng TH
                    if 'TH' not in loai_phong1.upper() and 'MÁY' not in loai_phong1.upper() and 'LAB' not in loai_phong1.upper():
                        warnings.append(f"⚠️ Lớp {tkb1.ma_lop.ma_lop} (TH) đang dùng phòng {tkb1.ma_phong.ma_phong} ({loai_phong1 or 'không xác định'})")
                else:
                    # Môn LT nên dùng phòng LT
                    if 'TH' in loai_phong1.upper() or 'MÁY' in loai_phong1.upper() or 'LAB' in loai_phong1.upper():
                        warnings.append(f"⚠️ Lớp {tkb1.ma_lop.ma_lop} (LT) đang dùng phòng {tkb1.ma_phong.ma_phong} ({loai_phong1})")
            
            if tkb2.ma_phong:
                mon2 = tkb2.ma_lop.ma_mon_hoc
                loai_phong2 = tkb2.ma_phong.loai_phong or ''
                
                if mon2.so_tiet_th and mon2.so_tiet_th > 0:
                    if 'TH' not in loai_phong2.upper() and 'MÁY' not in loai_phong2.upper() and 'LAB' not in loai_phong2.upper():
                        warnings.append(f"⚠️ Lớp {tkb2.ma_lop.ma_lop} (TH) đang dùng phòng {tkb2.ma_phong.ma_phong} ({loai_phong2 or 'không xác định'})")
                else:
                    if 'TH' in loai_phong2.upper() or 'MÁY' in loai_phong2.upper() or 'LAB' in loai_phong2.upper():
                        warnings.append(f"⚠️ Lớp {tkb2.ma_lop.ma_lop} (LT) đang dùng phòng {tkb2.ma_phong.ma_phong} ({loai_phong2})")
            
            # 3. Kiểm tra sức chứa phòng
            if tkb1.ma_phong and tkb1.ma_phong.suc_chua and tkb1.ma_lop.so_luong_sv:
                if tkb1.ma_lop.so_luong_sv > tkb1.ma_phong.suc_chua:
                    errors.append(f"❌ Lớp {tkb1.ma_lop.ma_lop} có {tkb1.ma_lop.so_luong_sv} SV nhưng phòng {tkb1.ma_phong.ma_phong} chỉ chứa {tkb1.ma_phong.suc_chua}")
            
            if tkb2.ma_phong and tkb2.ma_phong.suc_chua and tkb2.ma_lop.so_luong_sv:
                if tkb2.ma_lop.so_luong_sv > tkb2.ma_phong.suc_chua:
                    errors.append(f"❌ Lớp {tkb2.ma_lop.ma_lop} có {tkb2.ma_lop.so_luong_sv} SV nhưng phòng {tkb2.ma_phong.ma_phong} chỉ chứa {tkb2.ma_phong.suc_chua}")
        
        # Nếu không swap phòng, giữ nguyên phòng và validate
        else:
            # Kiểm tra phòng cũ có phù hợp với timeslot mới không
            if tkb1.ma_phong:
                conflict1 = ThoiKhoaBieu.objects.filter(
                    ma_dot=tkb1.ma_dot,
                    ma_phong=tkb1.ma_phong,
                    time_slot_id=tkb1.time_slot_id,
                    is_deleted=False
                ).exclude(ma_tkb=ma_tkb_1)
                
                if conflict1.exists():
                    errors.append(f"❌ Phòng {tkb1.ma_phong.ma_phong} đã bị trùng tại timeslot mới {tkb1.time_slot_id}")
            
            if tkb2.ma_phong:
                conflict2 = ThoiKhoaBieu.objects.filter(
                    ma_dot=tkb2.ma_dot,
                    ma_phong=tkb2.ma_phong,
                    time_slot_id=tkb2.time_slot_id,
                    is_deleted=False
                ).exclude(ma_tkb=ma_tkb_2)
                
                if conflict2.exists():
                    errors.append(f"❌ Phòng {tkb2.ma_phong.ma_phong} đã bị trùng tại timeslot mới {tkb2.time_slot_id}")
        
        # Validate ràng buộc GV (không được trùng timeslot)
        if gv1 and gv1.ma_gv:
            other_classes = PhanCong.objects.filter(
                ma_dot=tkb1.ma_dot,
                ma_gv=gv1.ma_gv
            ).exclude(ma_lop=tkb1.ma_lop).values_list('ma_lop__ma_lop', flat=True)
            
            conflict_gv1 = ThoiKhoaBieu.objects.filter(
                ma_dot=tkb1.ma_dot,
                ma_lop__ma_lop__in=other_classes,
                time_slot_id=tkb1.time_slot_id,
                is_deleted=False
            )
            
            if conflict_gv1.exists():
                errors.append(f"❌ GV {gv1.ma_gv.ten_gv} đã có lịch dạy tại {tkb1.time_slot_id}")
        
        if gv2 and gv2.ma_gv:
            other_classes = PhanCong.objects.filter(
                ma_dot=tkb2.ma_dot,
                ma_gv=gv2.ma_gv
            ).exclude(ma_lop=tkb2.ma_lop).values_list('ma_lop__ma_lop', flat=True)
            
            conflict_gv2 = ThoiKhoaBieu.objects.filter(
                ma_dot=tkb2.ma_dot,
                ma_lop__ma_lop__in=other_classes,
                time_slot_id=tkb2.time_slot_id,
                is_deleted=False
            )
            
            if conflict_gv2.exists():
                errors.append(f"❌ GV {gv2.ma_gv.ten_gv} đã có lịch dạy tại {tkb2.time_slot_id}")
        
        # Nếu có lỗi, trả về lỗi
        if errors:
            return JsonResponse({
                'status': 'error',
                'message': 'Không thể hoán đổi vì vi phạm ràng buộc',
                'errors': errors,
                'warnings': warnings
            }, status=400)
        
        # Lưu
        tkb1.save()
        tkb2.save()
        
        # Log
        user = request.user.username if request.user.is_authenticated else 'anonymous'
        swap_type = "cùng GV" if same_teacher else "khác GV"
        swap_room_str = "có swap phòng" if swap_phong else "giữ nguyên phòng"
        
        TKBLog.objects.create(
            ma_tkb=ma_tkb_1,
            action='SWAP',
            user=user,
            old_data=old_data_1,
            new_data={
                'ma_phong': tkb1.ma_phong.ma_phong if tkb1.ma_phong else None,
                'time_slot_id': tkb1.time_slot_id.time_slot_id,
                'swap_type': swap_type,
                'swap_phong': swap_phong
            },
            reason=f'Hoán đổi với {ma_tkb_2} ({swap_type}, {swap_room_str})'
        )
        
        TKBLog.objects.create(
            ma_tkb=ma_tkb_2,
            action='SWAP',
            user=user,
            old_data=old_data_2,
            new_data={
                'ma_phong': tkb2.ma_phong.ma_phong if tkb2.ma_phong else None,
                'time_slot_id': tkb2.time_slot_id.time_slot_id,
                'swap_type': swap_type,
                'swap_phong': swap_phong
            },
            reason=f'Hoán đổi với {ma_tkb_1} ({swap_type}, {swap_room_str})'
        )
        
        message = f'Hoán đổi lịch thành công ({swap_type}, {swap_room_str})'
        if warnings:
            message += f' với {len(warnings)} cảnh báo'
        
        return JsonResponse({
            'status': 'success',
            'message': message,
            'warnings': warnings,
            'swap_info': {
                'same_teacher': same_teacher,
                'swap_phong': swap_phong,
                'tkb1': {
                    'lop': tkb1.ma_lop.ma_lop,
                    'phong': tkb1.ma_phong.ma_phong if tkb1.ma_phong else None,
                    'timeslot': str(tkb1.time_slot_id)
                },
                'tkb2': {
                    'lop': tkb2.ma_lop.ma_lop,
                    'phong': tkb2.ma_phong.ma_phong if tkb2.ma_phong else None,
                    'timeslot': str(tkb2.time_slot_id)
                }
            }
        })
        
    except ThoiKhoaBieu.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Không tìm thấy một trong hai lịch học'
        }, status=404)
    except Exception as e:
        logger.exception(f"Lỗi khi hoán đổi TKB: {e}")
        return JsonResponse({
            'status': 'error',
            'message': f'Lỗi: {str(e)}'
        }, status=500)


@require_role('admin', 'truong_khoa')
def tkb_manage_view(request, ma_gv=None):
    """
    Trang quản lý TKB với layout 2 cột - Admin và Trưởng Khoa (chỉ khoa mình)
    
    URL Parameters:
        ma_gv: Mã giảng viên từ URL (dùng cho Trưởng Khoa)
    
    Phân quyền:
        - Admin: Quản lý toàn bộ TKB
        - Trưởng Khoa: Quản lý TKB khoa mình (filter theo ma_khoa)
    """
    # Lấy thông tin phân quyền
    role_info = get_user_role_info(request.user)
    user_role = role_info['role']
    
    # Redirect non-admin users từ URL cũ sang URL mới
    if user_role != 'admin' and not ma_gv:
        # Nếu user truy cập URL cũ (/admin/sap_lich/tkb-manage/), redirect sang URL mới
        ma_gv_current = role_info['ma_gv'] or request.user.username
        if user_role == 'truong_khoa':
            return redirect(f'/truong-khoa/{ma_gv_current}/quan-ly-tkb/')
        else:
            # Các role khác không được quản lý TKB
            return redirect(f'/giang-vien/{ma_gv_current}/xem-tkb/')
    
    # Validate ma_gv trong URL với user hiện tại
    if ma_gv:
        # Nếu không phải admin, phải check ma_gv khớp với user
        if user_role == 'truong_khoa':
            if ma_gv != role_info['ma_gv']:
                return HttpResponseForbidden("Bạn không có quyền quản lý TKB của người khác")
        # Các role khác không được truy cập
        elif user_role != 'admin':
            return HttpResponseForbidden("Bạn không có quyền quản lý TKB")
    
    ma_dot = request.GET.get('ma_dot', '')
    ma_khoa = request.GET.get('ma_khoa', '')
    view_type = request.GET.get('view_type', 'teacher')  # Default là 'teacher'
    selected_id = request.GET.get('selected_id', '')  # ma_gv hoặc ma_phong
    
    # Áp dụng ràng buộc theo phân quyền
    if user_role == 'truong_khoa':
        ma_khoa = role_info['ma_khoa']
    elif user_role == 'truong_bo_mon':
        ma_khoa = role_info['ma_khoa']
    elif user_role == 'giang_vien':
        # Giáo viên không được phép truy cập trang quản lý
        from django.contrib import messages
        messages.error(request, 'Bạn không có quyền truy cập trang quản lý TKB')
        # Redirect về trang xem TKB của giảng viên
        if role_info['ma_gv']:
            return redirect(f"/giang-vien/{role_info['ma_gv']}/xem-tkb/")
        return redirect('/admin/')
    
    from apps.scheduling.models import Khoa
    
    dot_list = DotXep.objects.all().order_by('-ma_dot')
    
    # Lấy danh sách khoa (filter theo quyền)
    if user_role == 'truong_khoa':
        khoa_list = Khoa.objects.filter(ma_khoa=role_info['ma_khoa']).order_by('ma_khoa')
    elif user_role == 'truong_bo_mon' and role_info['ma_khoa']:
        khoa_list = Khoa.objects.filter(ma_khoa=role_info['ma_khoa']).order_by('ma_khoa')
    else:
        khoa_list = Khoa.objects.all().order_by('ma_khoa')
    
    if not ma_dot and dot_list.exists():
        ma_dot = dot_list.first().ma_dot
    
    context = {
        **admin.site.each_context(request),
        'title': 'Quản lý Thời Khóa Biểu',
        'ma_dot': ma_dot,
        'ma_khoa': ma_khoa,
        'view_type': view_type,
        'selected_id': selected_id,
        'dots': dot_list,
        'khoa_list': khoa_list,
        'user_role': user_role,
        'role_info': role_info,
        'app_label': 'sap_lich',
        'opts': {
            'app_label': 'sap_lich',
            'model_name': 'tkbmanage',
            'verbose_name_plural': 'Quản lý TKB',
        },
        'segment': ['quan-ly-tkb'],
    }
    
    if ma_dot:
        try:
            dot_xep = DotXep.objects.get(ma_dot=ma_dot)
            context['dot_xep'] = dot_xep
            
            # Lấy danh sách lớp (filter theo khoa nếu có)
            lop_query = LopMonHoc.objects.filter(
                phan_cong_list__ma_dot=dot_xep
            ).select_related('ma_mon_hoc')
            
            if ma_khoa:
                lop_query = lop_query.filter(
                    phan_cong_list__ma_gv__ma_bo_mon__ma_khoa__ma_khoa=ma_khoa
                )
            
            # Thêm filter theo bộ môn nếu là Trưởng Bộ Môn
            if user_role == 'truong_bo_mon' and role_info['ma_bo_mon']:
                lop_query = lop_query.filter(
                    phan_cong_list__ma_gv__ma_bo_mon__ma_bo_mon=role_info['ma_bo_mon']
                )
            
            lop_list = lop_query.distinct()
            context['lops'] = lop_list
            
            # Lấy danh sách giáo viên (có lịch dạy trong đợt này)
            gv_query = PhanCong.objects.filter(
                ma_dot=dot_xep,
                ma_gv__isnull=False
            ).select_related('ma_gv', 'ma_gv__ma_bo_mon', 'ma_gv__ma_bo_mon__ma_khoa')
            
            if ma_khoa:
                gv_query = gv_query.filter(ma_gv__ma_bo_mon__ma_khoa__ma_khoa=ma_khoa)
            
            if user_role == 'truong_bo_mon' and role_info['ma_bo_mon']:
                gv_query = gv_query.filter(ma_gv__ma_bo_mon__ma_bo_mon=role_info['ma_bo_mon'])
            
            gv_list = gv_query.distinct().values_list('ma_gv__ma_gv', 'ma_gv__ten_gv').order_by('ma_gv__ma_gv')
            context['teachers'] = [{'ma_gv': gv[0], 'ten_gv': gv[1]} for gv in gv_list]
            
            # Lấy danh sách môn học (để tạo lớp mới)
            mon_hoc_query = MonHoc.objects.all()
            if ma_khoa:
                # Filter môn học theo khoa (qua bộ môn -> GV -> môn dạy)
                mon_hoc_query = mon_hoc_query.filter(
                    gv_day_list__ma_gv__ma_bo_mon__ma_khoa__ma_khoa=ma_khoa
                ).distinct()
            context['mon_hoc_list'] = mon_hoc_query.order_by('ma_mon_hoc')
            
            # Lấy danh sách phòng (tất cả phòng, kèm loại phòng)
            phong_query = PhongHoc.objects.all()
            phong_list = phong_query.order_by('ma_phong').values('ma_phong', 'suc_chua', 'loai_phong')
            context['phongs'] = list(phong_list)
            
            # Lấy danh sách timeslots
            timeslot_list = TimeSlot.objects.all().order_by('thu', 'ca')
            context['timeslots'] = timeslot_list
            
        except DotXep.DoesNotExist:
            context['error'] = f'Không tìm thấy đợt {ma_dot}'
    
    return render(request, 'admin/tkb_manage.html', context)


@csrf_exempt
@require_http_methods(["GET"])
def tkb_mini_schedule_api(request):
    """API lấy mini schedule cho trang quản lý"""
    try:
        ma_dot = request.GET.get('ma_dot', '')
        ma_khoa = request.GET.get('ma_khoa', '')
        view_type = request.GET.get('view_type', '')  # 'teacher' hoặc 'room'
        selected_id = request.GET.get('selected_id', '')  # ma_gv hoặc ma_phong
        
        if not ma_dot:
            dot_xep = DotXep.objects.order_by('-ngay_tao').first()
            if dot_xep:
                ma_dot = dot_xep.ma_dot
        else:
            dot_xep = DotXep.objects.get(ma_dot=ma_dot)
        
        if not dot_xep:
            return JsonResponse({
                'status': 'error',
                'message': 'Không tìm thấy đợt xếp lịch'
            }, status=404)
        
        # Lấy TKB hiện tại (chưa xóa)
        tkb_query = ThoiKhoaBieu.objects.filter(
            ma_dot=dot_xep,
            is_deleted=False
        ).select_related(
            'ma_lop', 'ma_lop__ma_mon_hoc', 'ma_phong',
            'time_slot_id', 'time_slot_id__ca'
        )
        
        # Chỉ hiển thị khi đã chọn GV hoặc phòng cụ thể
        if view_type == 'teacher' and selected_id:
            # Lấy các lớp mà GV dạy
            lop_gv = PhanCong.objects.filter(
                ma_dot=dot_xep, ma_gv__ma_gv=selected_id
            ).values_list('ma_lop__ma_lop', flat=True)
            tkb_query = tkb_query.filter(ma_lop__ma_lop__in=lop_gv)
        elif view_type == 'room' and selected_id:
            # Lấy TKB của phòng
            tkb_query = tkb_query.filter(ma_phong__ma_phong=selected_id)
        else:
            # Nếu chưa chọn GV hoặc phòng, trả về empty
            tkb_query = tkb_query.none()
        
        tkb_list = tkb_query.order_by('time_slot_id__thu', 'time_slot_id__ca')
        
        schedule = []
        for tkb in tkb_list:
            # Lấy tên GV từ PhanCong
            phan_cong = PhanCong.objects.filter(
                ma_dot=dot_xep,
                ma_lop=tkb.ma_lop
            ).select_related('ma_gv').first()
            
            ten_gv = phan_cong.ma_gv.ten_gv if phan_cong and phan_cong.ma_gv else 'N/A'
            ma_gv = phan_cong.ma_gv.ma_gv if phan_cong and phan_cong.ma_gv else None
            
            schedule.append({
                'ma_tkb': tkb.ma_tkb,
                'ma_lop': tkb.ma_lop.ma_lop,
                'mon_hoc': tkb.ma_lop.ma_mon_hoc.ten_mon_hoc,
                'ma_mon': tkb.ma_lop.ma_mon_hoc.ma_mon_hoc,
                'ma_phong': tkb.ma_phong.ma_phong if tkb.ma_phong else None,
                'thu': tkb.time_slot_id.thu,
                'ca': tkb.time_slot_id.ca.ma_khung_gio,
                'ten_ca': tkb.time_slot_id.ca.ten_ca,
                'time_slot_id': tkb.time_slot_id.time_slot_id,
                'tuan_hoc': tkb.tuan_hoc or '',
                'ten_gv': ten_gv,
                'ma_gv': ma_gv,
            })
        
        # Lấy TKB đã xóa (cũng filter theo khoa nếu có)
        deleted_query = ThoiKhoaBieu.objects.filter(
            ma_dot=dot_xep,
            is_deleted=True
        ).select_related(
            'ma_lop', 'ma_phong', 'time_slot_id'
        )
        
        if ma_khoa:
            deleted_query = deleted_query.filter(
                ma_lop__phan_cong_list__ma_gv__ma_bo_mon__ma_khoa__ma_khoa=ma_khoa
            ).distinct()
        
        deleted_list = deleted_query.order_by('-ngay_tao')[:20]
        
        deleted = []
        for tkb in deleted_list:
            deleted.append({
                'ma_tkb': tkb.ma_tkb,
                'ma_lop': tkb.ma_lop.ma_lop,
                'mon_hoc': tkb.ma_lop.ma_mon_hoc.ten_mon_hoc,
                'ma_phong': tkb.ma_phong.ma_phong if tkb.ma_phong else 'N/A',
                'thu': tkb.time_slot_id.thu,
                'ca': tkb.time_slot_id.ca.ma_khung_gio,
            })
        
        return JsonResponse({
            'status': 'success',
            'schedule': schedule,
            'deleted': deleted
        })
        
    except Exception as e:
        logger.exception(f"Lỗi khi lấy mini schedule: {e}")
        return JsonResponse({
            'status': 'error',
            'message': f'Lỗi: {str(e)}'
        }, status=500)



def build_mini_schedule(tkb_list):
    """Xây dựng mini schedule cho sidebar"""
    schedule = {}
    for thu in range(2, 9):
        schedule[thu] = {}
    
    for tkb in tkb_list:
        thu = tkb.time_slot_id.thu
        ca = tkb.time_slot_id.ca.ma_khung_gio
        
        if ca not in schedule[thu]:
            schedule[thu][ca] = []
        
        schedule[thu][ca].append({
            'ma_tkb': tkb.ma_tkb,
            'ma_lop': tkb.ma_lop.ma_lop,
            'mon_hoc': tkb.ma_lop.ma_mon_hoc.ten_mon_hoc,
            'phong': tkb.ma_phong.ma_phong if tkb.ma_phong else 'TBA'
        })
    
    return schedule


@csrf_exempt
@require_http_methods(["GET"])
def tkb_occupied_rooms_api(request):
    """API lấy danh sách phòng đã được sử dụng cho 1 timeslot"""
    try:
        ma_dot = request.GET.get('ma_dot', '')
        time_slot_id = request.GET.get('time_slot_id', '')
        
        if not ma_dot or not time_slot_id:
            return JsonResponse({
                'status': 'error',
                'message': 'Thiếu tham số ma_dot hoặc time_slot_id'
            }, status=400)
        
        # Lấy tất cả phòng đang được sử dụng trong timeslot này (không xóa)
        occupied_rooms = ThoiKhoaBieu.objects.filter(
            ma_dot__ma_dot=ma_dot,
            time_slot_id__time_slot_id=time_slot_id,
            is_deleted=False,
            ma_phong__isnull=False
        ).values_list('ma_phong__ma_phong', flat=True).distinct()
        
        return JsonResponse({
            'status': 'success',
            'occupied_rooms': list(occupied_rooms),
            'count': len(occupied_rooms)
        })
        
    except Exception as e:
        logger.exception(f"Lỗi khi lấy occupied rooms: {e}")
        return JsonResponse({
            'status': 'error',
            'message': f'Lỗi: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def tkb_mon_hoc_info_api(request):
    """API lấy thông tin môn học và các nhóm/tổ đã sử dụng"""
    try:
        ma_mon_hoc = request.GET.get('ma_mon_hoc', '')
        
        if not ma_mon_hoc:
            return JsonResponse({
                'status': 'error',
                'message': 'Thiếu tham số ma_mon_hoc'
            }, status=400)
        
        # Lấy thông tin môn học
        mon_hoc = MonHoc.objects.get(ma_mon_hoc=ma_mon_hoc)
        
        # Lấy tất cả lớp môn học đã tồn tại cho môn này
        existing_classes = LopMonHoc.objects.filter(
            ma_mon_hoc=mon_hoc
        ).values('nhom_mh', 'to_mh', 'ma_lop').order_by('nhom_mh', 'to_mh')
        
        # Phân tích nhóm đã dùng (1-10)
        used_groups = {}
        for lop in existing_classes:
            nhom = lop['nhom_mh']
            to = lop['to_mh']
            
            if nhom not in used_groups:
                used_groups[nhom] = []
            used_groups[nhom].append({
                'to': to,
                'ma_lop': lop['ma_lop']
            })
        
        # Tìm nhóm trống (1-10)
        available_groups = []
        for nhom in range(1, 11):
            if nhom not in used_groups:
                available_groups.append({
                    'nhom': nhom,
                    'status': 'empty',
                    'can_add_to': []
                })
            else:
                # Kiểm tra nhóm này còn chỗ cho tổ mới không
                existing_tos = [item['to'] for item in used_groups[nhom]]
                
                # Nếu là môn TH, mỗi nhóm tối đa 2 tổ (to=1, to=2)
                if mon_hoc.so_tiet_th and mon_hoc.so_tiet_th > 0:
                    can_add = []
                    if 1 not in existing_tos:
                        can_add.append(1)
                    if 2 not in existing_tos:
                        can_add.append(2)
                    
                    if can_add:
                        available_groups.append({
                            'nhom': nhom,
                            'status': 'partial',
                            'existing': used_groups[nhom],
                            'can_add_to': can_add
                        })
                else:
                    # Môn LT: mỗi nhóm chỉ có 1 lớp (to=0)
                    # Nếu đã có to=0 thì không thêm được nữa
                    pass
        
        # Suggest nhóm tiếp theo nên dùng
        suggested_group = None
        suggested_to = 0
        
        if mon_hoc.so_tiet_th and mon_hoc.so_tiet_th > 0:
            # Môn TH: Ưu tiên dùng hết tổ của nhóm cũ trước khi tạo nhóm mới
            for item in available_groups:
                if item['status'] == 'partial' and item['can_add_to']:
                    suggested_group = item['nhom']
                    suggested_to = item['can_add_to'][0]
                    break
            
            if not suggested_group:
                # Không có nhóm nào còn chỗ, tạo nhóm mới
                for item in available_groups:
                    if item['status'] == 'empty':
                        suggested_group = item['nhom']
                        suggested_to = 1
                        break
        else:
            # Môn LT: Tổ = 0, tìm nhóm trống đầu tiên
            for item in available_groups:
                if item['status'] == 'empty':
                    suggested_group = item['nhom']
                    suggested_to = 0
                    break
        
        return JsonResponse({
            'status': 'success',
            'mon_hoc': {
                'ma_mon_hoc': mon_hoc.ma_mon_hoc,
                'ten_mon_hoc': mon_hoc.ten_mon_hoc,
                'so_tiet_lt': mon_hoc.so_tiet_lt or 0,
                'so_tiet_th': mon_hoc.so_tiet_th or 0,
                'loai': 'TH' if (mon_hoc.so_tiet_th and mon_hoc.so_tiet_th > 0) else 'LT'
            },
            'used_groups': used_groups,
            'available_groups': available_groups,
            'suggested': {
                'nhom': suggested_group,
                'to': suggested_to
            }
        })
        
    except MonHoc.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': f'Không tìm thấy môn học: {ma_mon_hoc}'
        }, status=404)
    except Exception as e:
        logger.exception(f"Lỗi khi lấy thông tin môn học: {e}")
        return JsonResponse({
            'status': 'error',
            'message': f'Lỗi: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def tkb_gv_list_api(request):
    """API lấy danh sách giáo viên có lịch trong đợt xếp"""
    try:
        ma_dot = request.GET.get('ma_dot', '')
        
        if not ma_dot:
            return JsonResponse({
                'status': 'error',
                'message': 'Thiếu mã đợt xếp lịch'
            }, status=400)
        
        dot_xep = DotXep.objects.get(ma_dot=ma_dot)
        
        # Lấy danh sách GV có phân công trong đợt này
        gv_list = PhanCong.objects.filter(
            ma_dot=dot_xep,
            ma_gv__isnull=False
        ).select_related('ma_gv', 'ma_gv__ma_bo_mon').values(
            'ma_gv__ma_gv',
            'ma_gv__ten_gv',
            'ma_gv__ma_bo_mon__ten_bo_mon'
        ).distinct()
        
        # Chuyển thành list để sort theo TÊN (không phải họ)
        teachers = [
            {
                'ma_gv': gv['ma_gv__ma_gv'],
                'ten_gv': gv['ma_gv__ten_gv'],
                'bo_mon': gv['ma_gv__ma_bo_mon__ten_bo_mon'] or 'N/A'
            }
            for gv in gv_list
        ]
        
        # Sort theo TÊN (lấy từ cuối cùng của họ tên)
        # VD: "Nguyễn Văn A" → sort theo "A", "Trần Thị Bích" → sort theo "Bích"
        def get_first_name(full_name):
            parts = full_name.strip().split()
            return parts[-1] if parts else full_name
        
        teachers.sort(key=lambda x: get_first_name(x['ten_gv']))
        
        return JsonResponse({
            'status': 'success',
            'teachers': teachers
        })
        
    except DotXep.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Không tìm thấy đợt xếp lịch'
        }, status=404)
    except Exception as e:
        logger.exception(f"Lỗi khi lấy danh sách GV: {e}")
        return JsonResponse({
            'status': 'error',
            'message': f'Lỗi: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def tkb_gv_schedule_api(request):
    """API lấy thời khóa biểu của 1 giáo viên"""
    try:
        ma_dot = request.GET.get('ma_dot', '')
        ma_gv = request.GET.get('ma_gv', '')
        
        if not ma_dot or not ma_gv:
            return JsonResponse({
                'status': 'error',
                'message': 'Thiếu mã đợt xếp lịch hoặc mã giáo viên'
            }, status=400)
        
        dot_xep = DotXep.objects.get(ma_dot=ma_dot)
        
        # Lấy các lớp mà GV dạy
        lop_gv = PhanCong.objects.filter(
            ma_dot=dot_xep,
            ma_gv__ma_gv=ma_gv
        ).values_list('ma_lop__ma_lop', flat=True)
        
        # Lấy TKB của các lớp đó
        tkb_list = ThoiKhoaBieu.objects.filter(
            ma_dot=dot_xep,
            ma_lop__ma_lop__in=lop_gv,
            is_deleted=False
        ).select_related(
            'ma_lop', 'ma_lop__ma_mon_hoc',
            'ma_phong', 'time_slot_id', 'time_slot_id__ca'
        ).order_by('time_slot_id__thu', 'time_slot_id__ca')
        
        schedule = []
        for tkb in tkb_list:
            schedule.append({
                'ma_tkb': tkb.ma_tkb,
                'ma_lop': tkb.ma_lop.ma_lop,
                'mon_hoc': tkb.ma_lop.ma_mon_hoc.ten_mon_hoc,
                'ma_phong': tkb.ma_phong.ma_phong if tkb.ma_phong else None,
                'thu': tkb.time_slot_id.thu,
                'ca': tkb.time_slot_id.ca.ma_khung_gio,
                'time_slot_id': tkb.time_slot_id.time_slot_id,
                'timeslot_str': f"Thứ {tkb.time_slot_id.thu}, Ca {tkb.time_slot_id.ca.ma_khung_gio}"
            })
        
        return JsonResponse({
            'status': 'success',
            'schedule': schedule
        })
        
    except DotXep.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Không tìm thấy đợt xếp lịch'
        }, status=404)
    except Exception as e:
        logger.exception(f"Lỗi khi lấy lịch GV: {e}")
        return JsonResponse({
            'status': 'error',
            'message': f'Lỗi: {str(e)}'
        }, status=500)

"""
Excel Export Utility for Scheduling System
Xuất dữ liệu các bảng ra file Excel
"""

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


class ExcelExporter:
    """Utility class to export Django QuerySets to Excel"""
    
    @staticmethod
    def export_queryset(queryset, fields, headers, filename, title=None):
        """
        Export Django QuerySet to Excel file
        
        Args:
            queryset: Django QuerySet to export
            fields: List of field names to export (can use dot notation for related fields)
            headers: List of column headers
            filename: Name of the Excel file (without .xlsx extension)
            title: Optional title for the worksheet
        
        Returns:
            HttpResponse with Excel file
        """
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = title or "Data"
        
        # Styling
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Write data
        for row_num, obj in enumerate(queryset, 2):
            for col_num, field in enumerate(fields, 1):
                # Handle nested fields (e.g., 'ma_khoa.ten_khoa')
                value = obj
                for field_part in field.split('.'):
                    if hasattr(value, field_part):
                        value = getattr(value, field_part)
                    else:
                        value = ''
                        break
                
                # Handle callable fields (e.g., methods)
                if callable(value):
                    value = value()
                
                # Format datetime
                if isinstance(value, datetime):
                    value = value.strftime('%d/%m/%Y %H:%M')
                
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = str(value) if value is not None else ''
                cell.border = border
                cell.alignment = Alignment(vertical='center')
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            
            # Check header length
            max_length = len(headers[col_num - 1])
            
            # Check data length (sample first 100 rows)
            for row_num in range(2, min(102, ws.max_row + 1)):
                cell_value = ws[f'{column_letter}{row_num}'].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Set column width (add some padding)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        
        wb.save(response)
        return response
    
    @staticmethod
    def export_khoa(queryset):
        """Export Khoa (Faculty) data"""
        fields = ['ma_khoa', 'ten_khoa']
        headers = ['Mã Khoa', 'Tên Khoa']
        filename = f'Danh_sach_Khoa_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        return ExcelExporter.export_queryset(queryset, fields, headers, filename, 'Danh sách Khoa')
    
    @staticmethod
    def export_bo_mon(queryset):
        """Export BoMon (Department) data"""
        fields = ['ma_bo_mon', 'ten_bo_mon', 'ma_khoa.ten_khoa']
        headers = ['Mã Bộ Môn', 'Tên Bộ Môn', 'Khoa']
        filename = f'Danh_sach_Bo_Mon_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        return ExcelExporter.export_queryset(queryset, fields, headers, filename, 'Danh sách Bộ môn')
    
    @staticmethod
    def export_giang_vien(queryset):
        """Export GiangVien (Teacher) data"""
        fields = ['ma_gv', 'ten_gv', 'ma_bo_mon.ten_bo_mon', 'ma_bo_mon.ma_khoa.ten_khoa', 'loai_gv', 'email', 'ghi_chu']
        headers = ['Mã GV', 'Tên Giảng Viên', 'Bộ Môn', 'Khoa', 'Loại GV', 'Email', 'Ghi Chú']
        filename = f'Danh_sach_Giang_Vien_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        return ExcelExporter.export_queryset(queryset, fields, headers, filename, 'Danh sách Giảng viên')
    
    @staticmethod
    def export_mon_hoc(queryset):
        """Export MonHoc (Course) data"""
        fields = ['ma_mon_hoc', 'ten_mon_hoc', 'so_tin_chi', 'ly_thuyet', 'thuc_hanh', 'bai_tap']
        headers = ['Mã Môn', 'Tên Môn Học', 'Tín Chỉ', 'LT', 'TH', 'BT']
        filename = f'Danh_sach_Mon_Hoc_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        return ExcelExporter.export_queryset(queryset, fields, headers, filename, 'Danh sách Môn học')
    
    @staticmethod
    def export_phong_hoc(queryset):
        """Export PhongHoc (Room) data"""
        fields = ['ma_phong', 'ten_phong', 'loai_phong', 'suc_chua', 'co_so', 'ghi_chu']
        headers = ['Mã Phòng', 'Tên Phòng', 'Loại Phòng', 'Sức Chứa', 'Cơ Sở', 'Ghi Chú']
        filename = f'Danh_sach_Phong_Hoc_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        return ExcelExporter.export_queryset(queryset, fields, headers, filename, 'Danh sách Phòng học')
    
    @staticmethod
    def export_lop_mon_hoc(queryset):
        """Export LopMonHoc (Class) data"""
        fields = ['ma_lop_mon_hoc', 'ma_mon_hoc.ten_mon_hoc', 'loai_lop', 'si_so', 
                  'ma_du_kien_dt.ma_du_kien_dt', 'ghi_chu']
        headers = ['Mã Lớp MH', 'Môn Học', 'Loại Lớp', 'Sĩ Số', 'Dự Kiến ĐT', 'Ghi Chú']
        filename = f'Danh_sach_Lop_Mon_Hoc_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        return ExcelExporter.export_queryset(queryset, fields, headers, filename, 'Danh sách Lớp môn học')
    
    @staticmethod
    def export_phan_cong(queryset):
        """Export PhanCong (Assignment) data"""
        fields = ['ma_phan_cong', 'ma_lop_mon_hoc.ma_lop_mon_hoc', 
                  'ma_lop_mon_hoc.ma_mon_hoc.ten_mon_hoc', 'ma_gv.ten_gv', 
                  'ma_gv.ma_bo_mon.ten_bo_mon', 'so_tiet']
        headers = ['Mã Phân Công', 'Mã Lớp MH', 'Môn Học', 'Giảng Viên', 'Bộ Môn', 'Số Tiết']
        filename = f'Danh_sach_Phan_Cong_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        return ExcelExporter.export_queryset(queryset, fields, headers, filename, 'Danh sách Phân công')
    
    @staticmethod
    def export_thoi_khoa_bieu(queryset):
        """Export ThoiKhoaBieu (Schedule) data"""
        fields = ['ma_tkb', 'ma_phan_cong.ma_lop_mon_hoc.ma_lop_mon_hoc',
                  'ma_phan_cong.ma_lop_mon_hoc.ma_mon_hoc.ten_mon_hoc',
                  'ma_phan_cong.ma_gv.ten_gv', 'ma_phong.ten_phong',
                  'thu', 'tiet_bat_dau', 'so_tiet', 'ma_dot_xep.ten_dot_xep']
        headers = ['Mã TKB', 'Lớp MH', 'Môn Học', 'Giảng Viên', 'Phòng', 
                   'Thứ', 'Tiết BĐ', 'Số Tiết', 'Đợt Xếp']
        filename = f'Thoi_Khoa_Bieu_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        return ExcelExporter.export_queryset(queryset, fields, headers, filename, 'Thời khóa biểu')
    
    @staticmethod
    def export_nguyen_vong(queryset):
        """Export NguyenVong (Preference) data"""
        fields = ['ma_nguyen_vong', 'ma_gv.ten_gv', 'ma_gv.ma_bo_mon.ten_bo_mon',
                  'ma_dot_xep.ten_dot_xep', 'thu', 'tiet_bat_dau', 'so_tiet', 
                  'loai_nguyen_vong', 'ghi_chu']
        headers = ['Mã NV', 'Giảng Viên', 'Bộ Môn', 'Đợt Xếp', 'Thứ', 
                   'Tiết BĐ', 'Số Tiết', 'Loại NV', 'Ghi Chú']
        filename = f'Nguyen_Vong_GV_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        return ExcelExporter.export_queryset(queryset, fields, headers, filename, 'Nguyện vọng GV')
    
    @staticmethod
    def export_dot_xep(queryset):
        """Export DotXep (Schedule Period) data"""
        fields = ['ma_dot_xep', 'ten_dot_xep', 'ma_du_kien_dt.ma_du_kien_dt', 
                  'ngay_bat_dau', 'ngay_ket_thuc', 'trang_thai', 'ghi_chu']
        headers = ['Mã Đợt', 'Tên Đợt Xếp', 'Dự Kiến ĐT', 'Ngày BĐ', 
                   'Ngày KT', 'Trạng Thái', 'Ghi Chú']
        filename = f'Danh_sach_Dot_Xep_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        return ExcelExporter.export_queryset(queryset, fields, headers, filename, 'Danh sách Đợt xếp')
    
    @staticmethod
    def export_gvdaymon(queryset):
        """Export GVDayMon data"""
        fields = ['id', 'ma_gv.ma_gv', 'ma_gv.ten_gv', 'ma_mon_hoc.ma_mon_hoc', 'ma_mon_hoc.ten_mon_hoc']
        headers = ['ID', 'Mã GV', 'Tên Giảng Viên', 'Mã Môn', 'Tên Môn Học']
        filename = f'GV_Day_Mon_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        return ExcelExporter.export_queryset(queryset, fields, headers, filename, 'GV dạy môn')
    
    @staticmethod
    def export_dukiendt(queryset):
        """Export DuKienDT data"""
        fields = ['ma_du_kien_dt', 'nam_hoc', 'hoc_ky', 'mo_ta_hoc_ky']
        headers = ['Mã Dự Kiến ĐT', 'Năm Học', 'Học Kỳ', 'Mô Tả']
        filename = f'Du_Kien_DT_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        return ExcelExporter.export_queryset(queryset, fields, headers, filename, 'Dự kiến đào tạo')
    
    @staticmethod
    def export_khungtg(queryset):
        """Export KhungTG data"""
        fields = ['ma_khung_gio', 'ten_ca', 'gio_bat_dau', 'gio_ket_thuc', 'so_tiet']
        headers = ['Mã Khung Giờ', 'Tên Ca', 'Giờ BĐ', 'Giờ KT', 'Số Tiết']
        filename = f'Khung_Thoi_Gian_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        return ExcelExporter.export_queryset(queryset, fields, headers, filename, 'Khung thời gian')
    
    @staticmethod
    def export_rangbuocmem(queryset):
        """Export RangBuocMem data"""
        fields = ['ma_rang_buoc', 'ten_rang_buoc', 'trong_so', 'mo_ta']
        headers = ['Mã Ràng Buộc', 'Tên Ràng Buộc', 'Trọng Số', 'Mô Tả']
        filename = f'Rang_Buoc_Mem_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        return ExcelExporter.export_queryset(queryset, fields, headers, filename, 'Ràng buộc mềm')
    
    @staticmethod
    def export_rangbuoctrongdot(queryset):
        """Export RangBuocTrongDot data"""
        fields = ['id', 'ma_dot.ma_dot', 'ma_dot.ten_dot_xep', 'ma_rang_buoc.ma_rang_buoc', 'ma_rang_buoc.ten_rang_buoc']
        headers = ['ID', 'Mã Đợt', 'Tên Đợt', 'Mã Ràng Buộc', 'Tên Ràng Buộc']
        filename = f'Rang_Buoc_Trong_Dot_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        return ExcelExporter.export_queryset(queryset, fields, headers, filename, 'Ràng buộc trong đợt')
    
    @staticmethod
    def export_ngaynghicodinh(queryset):
        """Export NgayNghiCoDinh data"""
        fields = ['id', 'ten_ngay_nghi', 'ngay', 'ghi_chu']
        headers = ['ID', 'Tên Ngày Nghỉ', 'Ngày', 'Ghi Chú']
        filename = f'Ngay_Nghi_Co_Dinh_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        return ExcelExporter.export_queryset(queryset, fields, headers, filename, 'Ngày nghỉ cố định')
    
    @staticmethod
    def export_ngaynghidot(queryset):
        """Export NgayNghiDot data"""
        fields = ['id', 'ma_dot.ma_dot', 'ma_dot.ten_dot_xep', 'ten_ngay_nghi', 'ngay_bd', 'so_ngay_nghi', 'tuan_bd', 'tuan_kt', 'ghi_chu']
        headers = ['ID', 'Mã Đợt', 'Tên Đợt', 'Tên Ngày Nghỉ', 'Ngày BĐ', 'Số Ngày Nghỉ', 'Tuần BĐ', 'Tuần KT', 'Ghi Chú']
        filename = f'Ngay_Nghi_Dot_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        return ExcelExporter.export_queryset(queryset, fields, headers, filename, 'Ngày nghỉ theo đợt')
    
    @staticmethod
    def export_timeslot(queryset):
        """Export TimeSlot data"""
        fields = ['time_slot_id', 'thu', 'ca']
        headers = ['Time Slot ID', 'Thứ', 'Ca']
        filename = f'Time_Slot_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        return ExcelExporter.export_queryset(queryset, fields, headers, filename, 'Time slots')

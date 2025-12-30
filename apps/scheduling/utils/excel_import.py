"""
Excel Import Utility for Django Admin
Simple and efficient Excel import for scheduling models
"""

import pandas as pd
from io import BytesIO
import re
import unicodedata
from django.http import HttpResponse
from django.contrib import messages


class ExcelImporter:
    """Handle Excel import operations for Django models"""
    
    @staticmethod
    def generate_template(model_class):
        """Generate Excel template for a model"""
        field_info = []
        
        pk_field = model_class._meta.pk
        
        # Add primary key to template (for update scenarios)
        if not pk_field.auto_created:  # CharField PK like ma_khoa, ma_gv
            field_info.append({
                'name': pk_field.name,
                'verbose_name': pk_field.verbose_name if hasattr(pk_field, 'verbose_name') else pk_field.name,
                'required': True,
                'type': pk_field.get_internal_type(),
                'max_length': pk_field.max_length if hasattr(pk_field, 'max_length') else None,
                'is_pk': True
            })
        
        for field in model_class._meta.get_fields():
            # Skip auto primary key (id) and reverse relations
            if field.name == pk_field.name:
                continue
            if hasattr(field, 'related_model') and field.related_model and field.many_to_one:
                # Include ForeignKey fields but show the ID field
                field_info.append({
                    'name': field.name,
                    'verbose_name': field.verbose_name if hasattr(field, 'verbose_name') else field.name,
                    'required': not field.blank if hasattr(field, 'blank') else True,
                    'type': f"ForeignKey -> {field.related_model.__name__}",
                    'max_length': None,
                    'is_fk': True
                })
                continue
            if hasattr(field, 'remote_field') and field.remote_field:
                continue
                
            field_info.append({
                'name': field.name,
                'verbose_name': field.verbose_name if hasattr(field, 'verbose_name') else field.name,
                'required': not field.blank if hasattr(field, 'blank') else True,
                'type': field.get_internal_type() if hasattr(field, 'get_internal_type') else 'CharField',
                'max_length': field.max_length if hasattr(field, 'max_length') else None,
            })
        
        # Create Excel file
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheet 1: Template
            df_template = pd.DataFrame(columns=[f['verbose_name'] for f in field_info])
            df_template.to_excel(writer, sheet_name='Dữ liệu', index=False)
            
            # Sheet 2: Instructions
            instructions = pd.DataFrame({
                'Tên cột': [f['verbose_name'] for f in field_info],
                'Tên field': [f['name'] for f in field_info],
                'Bắt buộc': ['Có' if f['required'] else 'Không' for f in field_info],
                'Kiểu': [f['type'] for f in field_info],
                'Độ dài tối đa': [str(f['max_length']) if f['max_length'] else 'N/A' for f in field_info],
                'Ghi chú': ['ID chính (nhập để cập nhật, để trống để tạo mới)' if f.get('is_pk') 
                           else 'Nhập mã của bản ghi liên kết' if f.get('is_fk')
                           else '' for f in field_info]
            })
            instructions.to_excel(writer, sheet_name='Hướng dẫn', index=False)
            
            # Format the sheets
            workbook = writer.book
            
            # Format Data sheet
            worksheet = workbook['Dữ liệu']
            from openpyxl.styles import Font, PatternFill
            for idx, field in enumerate(field_info, start=1):
                cell = worksheet.cell(row=1, column=idx)
                if field.get('is_pk'):
                    # Primary key - blue bold
                    cell.font = Font(bold=True, color='0000FF')
                    cell.fill = PatternFill(start_color='E0F0FF', end_color='E0F0FF', fill_type='solid')
                elif field['required']:
                    # Required - red bold
                    cell.font = Font(bold=True, color='FF0000')
                else:
                    cell.font = Font(bold=True)
            
            # Format Instructions sheet
            worksheet_inst = workbook['Hướng dẫn']
            for cell in worksheet_inst[1]:
                cell.font = Font(bold=True)
        
        output.seek(0)
        
        model_name = model_class._meta.verbose_name_plural.replace(' ', '_')
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{model_name}_template.xlsx"'
        
        return response
    
    @staticmethod
    def validate_and_import(file, model_class, request):
        """Validate and import Excel data into database"""
        try:
            # Read Excel file
            df = pd.read_excel(file, sheet_name='Dữ liệu')

            # Normalize column headers: strip, lower, remove accents
            def normalize_header(h):
                if not isinstance(h, str):
                    return h
                h = unicodedata.normalize('NFKD', h).encode('ascii', 'ignore').decode('ascii')
                h = h.strip().lower()
                h = re.sub(r'\s+', ' ', h)  # collapse multiple spaces
                return h

            normalized_cols = {col: normalize_header(col) for col in df.columns}
            df.rename(columns=normalized_cols, inplace=True)
            
            if df.empty:
                messages.error(request, "File Excel không có dữ liệu")
                return False
            
            # Helper: normalized verbose name
            def norm_verbose(name):
                if not isinstance(name, str):
                    return name
                name = unicodedata.normalize('NFKD', name).encode('ascii', 'ignore').decode('ascii')
                name = name.strip().lower()
                name = re.sub(r'\s+', ' ', name)
                return name

            # Get model fields info
            pk_field = model_class._meta.pk
            field_mapping = {}  # normalized verbose_name -> field_name
            required_fields = []
            fk_fields = {}  # field_name -> related_model

            # Add primary key if not auto-created
            if not pk_field.auto_created:
                verbose_name = pk_field.verbose_name if hasattr(pk_field, 'verbose_name') else pk_field.name
                field_mapping[norm_verbose(verbose_name)] = pk_field.name
            
            for field in model_class._meta.get_fields():
                if field.name == pk_field.name:
                    continue
                    
                # Handle ForeignKey fields
                if hasattr(field, 'related_model') and field.related_model and field.many_to_one:
                    verbose_name = field.verbose_name if hasattr(field, 'verbose_name') else field.name
                    field_mapping[norm_verbose(verbose_name)] = field.name
                    fk_fields[field.name] = field.related_model
                    
                    if hasattr(field, 'blank') and not field.blank:
                        required_fields.append(norm_verbose(verbose_name))
                    continue
                    
                if hasattr(field, 'remote_field') and field.remote_field:
                    continue
                
                verbose_name = field.verbose_name if hasattr(field, 'verbose_name') else field.name
                field_mapping[norm_verbose(verbose_name)] = field.name
                
                if hasattr(field, 'blank') and not field.blank:
                    required_fields.append(norm_verbose(verbose_name))
            
            # Validate required columns (normalized)
            df_cols_norm = [col for col in df.columns]
            missing_cols = [col for col in required_fields if col not in df_cols_norm]
            if missing_cols:
                messages.error(request, f"Thiếu các cột bắt buộc: {', '.join(missing_cols)}")
                return False

            # Check for empty required fields
            for col in required_fields:
                if col in df.columns:
                    empty_rows = df[df[col].isna()].index.tolist()
                    if empty_rows:
                        messages.error(request, f"Cột '{col}' có giá trị trống ở dòng: {', '.join([str(r+2) for r in empty_rows[:5]])}")
                        return False
            
            # Import data
            created_count = 0
            updated_count = 0
            error_count = 0
            errors = []
            
            for idx, row in df.iterrows():
                try:
                    # Convert row to dict with field names
                    data = {}
                    fk_data = {}  # Separate FK data for lookup
                    
                    for verbose_name, field_name in field_mapping.items():
                        if verbose_name in df.columns and pd.notna(row[verbose_name]):
                            value = row[verbose_name]
                            
                            # Handle ForeignKey - store for later validation
                            if field_name in fk_fields:
                                fk_data[field_name] = value
                            else:
                                data[field_name] = value
                    
                    if not data and not fk_data:
                        continue
                    
                    # Get PK value from data
                    pk_value = data.get(pk_field.name)
                    
                    # Validate and resolve ForeignKeys
                    for fk_field, fk_value in fk_data.items():
                        related_model = fk_fields[fk_field]
                        related_pk = related_model._meta.pk.name
                        try:
                            related_obj = related_model.objects.get(**{related_pk: fk_value})
                            data[fk_field] = related_obj
                        except related_model.DoesNotExist:
                            raise ValueError(f"Không tìm thấy {related_model._meta.verbose_name} với mã '{fk_value}'")
                    
                    if pk_value:
                        # Check if exists
                        exists = model_class.objects.filter(**{pk_field.name: pk_value}).exists()
                        
                        if exists:
                            # Update existing record
                            update_data = {k: v for k, v in data.items() if k != pk_field.name}
                            if update_data:
                                model_class.objects.filter(**{pk_field.name: pk_value}).update(**update_data)
                                updated_count += 1
                        else:
                            # Create new with specified PK
                            model_class.objects.create(**data)
                            created_count += 1
                    else:
                        # Create new - let Django auto-generate PK if auto field
                        if pk_field.auto_created:
                            model_class.objects.create(**data)
                            created_count += 1
                        else:
                            raise ValueError(f"Thiếu giá trị cho trường khóa chính '{pk_field.verbose_name}'")
                        
                except Exception as e:
                    error_count += 1
                    if len(errors) < 10:  # Only store first 10 errors
                        errors.append(f"Dòng {idx+2}: {str(e)}")
            
            # Show results
            if created_count > 0:
                messages.success(request, f"Đã tạo mới {created_count} bản ghi")
            if updated_count > 0:
                messages.info(request, f"Đã cập nhật {updated_count} bản ghi")
            if error_count > 0:
                messages.warning(request, f"Có {error_count} lỗi")
                for error in errors:
                    messages.error(request, error)
            if created_count == 0 and updated_count == 0 and error_count == 0:
                messages.warning(request, "Không có bản ghi nào được import. Kiểm tra lại dữ liệu và tên cột.")
            
            return created_count > 0 or updated_count > 0
            
        except Exception as e:
            messages.error(request, f"Lỗi xử lý file: {str(e)}")
            return False
